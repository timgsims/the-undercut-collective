#!/usr/bin/env python3
"""
The Undercut Collective — F1 Fantasy Dashboard Builder
=======================================================
Reads your Excel workbook and generates a combined HTML dashboard,
then commits and pushes it to GitHub Pages.

Run this after every race:  python build_dashboard.py
Or just double-click it in File Explorer.

Requirements (install once):
    pip install openpyxl
"""

import sys
import os
import re
import subprocess
from pathlib import Path

# ── Try importing openpyxl, give clear instructions if missing ────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Fix:   Open a terminal and run:  pip install openpyxl")
    input("\nPress Enter to exit...")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — edit these if paths change
# ─────────────────────────────────────────────────────────────────────────────

WORKBOOK_PATH = r"C:\Users\tim\OneDrive\Documents\Excel\the-undercut-collective\The Undercut Collective F1 Fantasy League.xlsx"

# Folder where this script lives (= your GitHub repo root)
REPO_DIR = Path(__file__).parent

# The output HTML file (must be index.html for GitHub Pages)
OUTPUT_FILE = REPO_DIR / "index.html"

# Git settings — leave GITHUB_REMOTE blank to skip auto-push
GITHUB_REMOTE = "origin"   # set to "" to disable auto-push
COMMIT_MSG_PREFIX = "Update dashboard"  # race name appended automatically

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

SEASON = 2026

MANAGER_COLOURS = {
    "Tim":     "#378ADD",
    "Stu":     "#1D9E75",
    "Jaime":   "#D85A30",
    "Grayson": "#7F77DD",
    "Cain":    "#BA7517",
    "Lori":    "#D4537E",
    "Mark":    "#888780",
    "Dan":     "#5DCAA5",
    "Mike":    "#E24B4A",
}

CHIP_STYLES = {
    "Limitless":   {"bg": "#1a3a5c", "tc": "#60aaff"},
    "Wildcard":    {"bg": "#4a1a12", "tc": "#ff7a5a"},
    "Final Fix":   {"bg": "#3d2200", "tc": "#ffaa44"},
    "Auto Pilot":  {"bg": "#0d3328", "tc": "#3ddbb0"},
    "No Negative": {"bg": "#251a4a", "tc": "#b39dff"},
    "Extra DRS":   {"bg": "#1a3010", "tc": "#88dd44"},
}

CHIP_ORDER = ["Limitless", "Wildcard", "Final Fix", "Auto Pilot", "No Negative", "Extra DRS"]

DASH_PATTERNS = [[], [6,2], [2,2], [8,3], [4,2], [6,2,2,2], [3,3], [8,2,2,2], [1,2]]

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — READ EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def read_workbook(path):
    print(f"Reading workbook: {path}")
    if not Path(path).exists():
        print(f"\nERROR: Workbook not found at:\n  {path}")
        print("Check the WORKBOOK_PATH setting at the top of this script.")
        input("\nPress Enter to exit...")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # ── Races sheet (sheet index 0) ──────────────────────────────────────────
    ws = wb.worksheets[0]   # Races

    # Row 2 = headers. Even columns H onwards = race scores (col 8, 10, 12...)
    # Row 1 = round numbers in even cols.  Row 2 = race names in even cols.
    # Cols: A=Team, B=Name, C=Total, D=Rank, E=Interval, F=Gap, G=Avg, H onwards = scores

    # Build race list from row 2 (col 8 = H onwards, every 2)
    races = []       # list of {"round": int, "name": str, "col": int}
    for col in range(8, ws.max_column + 1, 2):
        name = ws.cell(row=2, column=col).value
        rnd  = ws.cell(row=1, column=col).value
        if name and rnd:
            races.append({"round": int(rnd), "name": str(name).strip(), "col": col})

    # Managers (rows 3–11)
    managers_raw = []
    for row in range(3, 12):
        name_cell = ws.cell(row=row, column=2).value
        if not name_cell:
            break
        name  = str(name_cell).strip()
        team  = str(ws.cell(row=row, column=1).value or "").strip()
        total = ws.cell(row=row, column=3).value or 0
        rank  = ws.cell(row=row, column=4).value or 0
        avg   = ws.cell(row=row, column=7).value or 0

        scores = {}
        for race in races:
            v = ws.cell(row=row, column=race["col"]).value
            if v is not None and v != "":
                try:
                    scores[race["name"]] = int(float(str(v)))
                except (ValueError, TypeError):
                    pass   # position string in odd col — skip

        managers_raw.append({
            "name":   name,
            "team":   team,
            "total":  int(float(str(total))) if total else 0,
            "rank":   int(rank) if rank else 0,
            "avg":    float(str(avg)) if avg else 0,
            "scores": scores,   # {race_name: pts}
            "color":  MANAGER_COLOURS.get(name, "#888888"),
        })

    # Sort by rank
    managers_raw.sort(key=lambda m: m["rank"])
    data["managers"]    = managers_raw
    data["races"]       = races
    data["races_done"]  = [r for r in races if any(
        m["scores"].get(r["name"]) is not None for m in managers_raw)]

    # Podiums are derived later in compute() from per-race score rankings —
    # same source Race Breakdown uses — instead of read from a worksheet, so the
    # two pages can never drift apart.

    # ── Stats sheet (index 2) — finish distribution + chip usage ────────────
    ws_stats = wb.worksheets[2]
    finish_dist = {}     # {name: [p1, p2, ..., pN]} — N detected dynamically
    chips_used  = {}     # {name: {chip_name: bool}}

    # Rows 2–10: finish distribution.
    # The number of finish-position columns equals the number of managers in the league —
    # one column per possible finishing position (P1 through P_N_managers).
    # This is more reliable than sniffing data types, which picks up the Podiums
    # and Total Pts columns that also contain integers.
    n_finish_cols = len(managers_raw)
    data["n_finish_cols"] = n_finish_cols

    for row in range(2, 12):
        name = ws_stats.cell(row=row, column=2).value
        if not name:
            break
        name = str(name).strip()
        finishes = []
        for col in range(3, 3 + n_finish_cols):
            v = ws_stats.cell(row=row, column=col).value
            finishes.append(int(float(str(v))) if v else 0)
        finish_dist[name] = finishes

    # Chip usage: header row 12, data rows 13+
    # Col B=Name, C=Limitless, D=Wildcard, E=Final Fix, F=Auto Pilot, G=No Negative, H=Extra DRS
    chip_header_row = None
    for row in range(11, 25):
        if ws_stats.cell(row=row, column=1).value == "CHIPS USED":
            chip_header_row = row
            break

    if chip_header_row:
        chip_cols = {}   # {col: chip_name}
        for col in range(3, 9):
            hdr = ws_stats.cell(row=chip_header_row, column=col).value
            if hdr:
                chip_cols[col] = str(hdr).strip()

        for row in range(chip_header_row + 1, chip_header_row + 12):
            name = ws_stats.cell(row=row, column=2).value
            if not name:
                continue
            name = str(name).strip()
            used = {}
            for col, chip in chip_cols.items():
                v = ws_stats.cell(row=row, column=col).value
                used[chip] = bool(v and str(v).strip().upper() == "X")
            chips_used[name] = used

    data["finish_dist"] = finish_dist
    data["chips_used"]  = chips_used

    # ── Stats sheet — chip usage BY RACE (col P onwards, rows 1–9) ───────────
    # Row 1: col P = "CHIPS USED BY RACE", col Q+ = race names
    # Rows 2–9: col P = manager name, col Q+ = chip name used that race (or empty)
    chip_race_usage = {}   # {manager_name: {race_name: chip_name}}
    chip_by_race_header_col = None
    for col in range(1, ws_stats.max_column + 1):   # scan whole row 1 for the label
        v = ws_stats.cell(row=1, column=col).value
        if v and "CHIPS USED BY RACE" in str(v):
            chip_by_race_header_col = col
            break

    if chip_by_race_header_col:
        # Read race names from row 1 starting one column after the label
        cbr_race_cols = {}   # {col: race_name}
        for col in range(chip_by_race_header_col + 1, ws_stats.max_column + 1):
            rname = ws_stats.cell(row=1, column=col).value
            if rname and str(rname).strip():
                cbr_race_cols[col] = str(rname).strip()
            else:
                break
        # Read manager rows (rows 2–9, col P = manager name)
        for row in range(2, 11):
            name = ws_stats.cell(row=row, column=chip_by_race_header_col).value
            if not name:
                continue
            name = str(name).strip()
            used_by_race = {}
            for col, rname in cbr_race_cols.items():
                v = ws_stats.cell(row=row, column=col).value
                if v and str(v).strip():
                    used_by_race[rname] = str(v).strip()
            chip_race_usage[name] = used_by_race

    data["chip_race_usage"] = chip_race_usage

    # ── Driver Changes sheet (index 3) — lineups + actual results ────────────
    ws_dc = wb.worksheets[3]

    # Layout (confirmed from XML):
    #   Row 1: round numbers in race-name cols (D=1, I=2, N=3, S=4 ...)
    #   Row 2: A=Team, B=Name, C=Starting Value, then per race (5-col blocks):
    #     [race_col+0] = race name   (e.g. D=Australia, I=China, N=Japan ...)
    #     [race_col+1] = DRS marker  ("2X" if DRS pick, else empty)
    #     [race_col+2] = Budget +-
    #     [race_col+3] = Value
    #     [race_col+4] = Result      (points manager earned for this pick this race)
    #   Race blocks start at col 4 (D=index 4 in 1-based) and repeat every 5 cols.
    #   Manager section: rows 3–58 (7 rows per manager, 5 drivers + 2 constructors)
    #   Actual driver results: rows 61–82 (same column layout, col E=2X doubled score)
    #   Constructor results:   rows 83–93 (same column layout, no DRS column)

    skip_hdrs = {"2X", "Budget +-", "Value", "#", "DRS", "", None}

    # Detect race columns from row 2 (1-based col 4 = D onward, every 5)
    race_cols_dc = []   # [(race_name, col_1based), ...]
    for col in range(4, ws_dc.max_column + 1, 5):
        hdr = ws_dc.cell(row=2, column=col).value
        if hdr and str(hdr).strip() not in skip_hdrs:
            race_cols_dc.append((str(hdr).strip(), col))

    # --- Manager lineup section (rows 3–58, 7 rows per manager) ---
    lineups = {}   # {manager_name: {race_name: [{"name", "drs", "pts", "is_constructor"}, ...]}}
    row = 3
    while row <= 58:
        team = ws_dc.cell(row=row, column=1).value
        name = ws_dc.cell(row=row, column=2).value
        if not team or not name:
            row += 7
            continue
        name = str(name).strip()
        if name not in lineups:
            lineups[name] = {}

        for race_name, col in race_cols_dc:
            picks = []
            for r in range(row, row + 7):
                pick_name  = ws_dc.cell(row=r, column=col).value
                drs_marker = ws_dc.cell(row=r, column=col + 1).value
                pts_val    = ws_dc.cell(row=r, column=col + 4).value
                if pick_name and str(pick_name).strip():
                    pname = str(pick_name).strip()
                    is_con = "." not in pname   # drivers have "X. Surname" initials
                    try:
                        pts = int(float(str(pts_val))) if pts_val not in (None, "", "#N/A") else None
                    except (ValueError, TypeError):
                        pts = None
                    picks.append({
                        "name":           pname,
                        "drs":            bool(drs_marker and str(drs_marker).strip() in ("2X", "3X")),
                        "drs_marker":     str(drs_marker).strip() if drs_marker and str(drs_marker).strip() in ("2X", "3X") else "",
                        "pts":            pts,   # already the final earned pts (doubled/tripled by formula)
                        "is_constructor": is_con,
                    })
            if picks:
                lineups[name][race_name] = picks

        row += 7

    data["lineups"] = lineups

    # --- Actual driver results (rows 61–82) ---
    # col+0 = base pts, col+1 = doubled pts (skip), col+2 = Budget +-, col+3 = Value
    driver_results  = {}   # {driver_name: {race_name: {"pts": int, "bud": float}}}
    for row in range(61, 83):
        dname = ws_dc.cell(row=row, column=1).value
        if not dname:
            continue
        dname = str(dname).strip()
        driver_results[dname] = {}
        for race_name, col in race_cols_dc:
            pts_v = ws_dc.cell(row=row, column=col).value
            bud_v = ws_dc.cell(row=row, column=col + 2).value
            try:
                pts = int(float(str(pts_v))) if pts_v not in (None, "", "#N/A") else None
            except (ValueError, TypeError):
                pts = None
            try:
                bud = round(float(str(bud_v)), 2) if bud_v not in (None, "", "#N/A") else None
            except (ValueError, TypeError):
                bud = None
            driver_results[dname][race_name] = {"pts": pts, "bud": bud}

    data["driver_results"] = driver_results

    # --- Actual constructor results (rows 83–93) ---
    # col+0 = pts, col+1 = Budget +- (no DRS column for constructors), col+2 = Value
    constructor_results = {}   # {constructor_name: {race_name: {"pts": int, "bud": float}}}
    for row in range(83, 94):
        bval = ws_dc.cell(row=row, column=2).value   # col B = name
        if not bval or str(bval).strip() in ("", "B", "Team"):
            continue
        cname = str(bval).strip()
        constructor_results[cname] = {}
        for race_name, col in race_cols_dc:
            pts_v = ws_dc.cell(row=row, column=col).value
            bud_v = ws_dc.cell(row=row, column=col + 1).value   # constructors: bud is col+1
            try:
                pts = int(float(str(pts_v))) if pts_v not in (None, "", "#N/A") else None
            except (ValueError, TypeError):
                pts = None
            try:
                bud = round(float(str(bud_v)), 2) if bud_v not in (None, "", "#N/A") else None
            except (ValueError, TypeError):
                bud = None
            constructor_results[cname][race_name] = {"pts": pts, "bud": bud}

    data["constructor_results"] = constructor_results

    # ── Reference Tables sheet (index 9) — budgets ──────────────────────────
    ws_ref = wb.worksheets[9]

    # Budget section starts at row 41 based on XML analysis
    # Row 41 = header: Name, Pre-Season, Australia, China, ...
    # Rows 42–50 = one per manager
    # Anchor: find the row with exactly "Budget" in col A (not "Budget Change")
    # then header row is immediately after
    budget_header_row = None
    for row in range(35, 55):
        cell = ws_ref.cell(row=row, column=1).value
        if cell and str(cell).strip() == "Budget":
            budget_header_row = row + 1
            break

    budgets = {}   # {name: [pre_season, r1, r2, ...]}
    budget_race_names = []

    if budget_header_row:
        # Read race headers from row
        for col in range(2, ws_ref.max_column + 1):
            hdr = ws_ref.cell(row=budget_header_row, column=col).value
            if hdr:
                budget_race_names.append(str(hdr).strip())
            else:
                break

        for row in range(budget_header_row + 1, budget_header_row + 12):
            name = ws_ref.cell(row=row, column=1).value
            if not name:
                continue
            name = str(name).strip()
            vals = []
            for col in range(2, 2 + len(budget_race_names)):
                v = ws_ref.cell(row=row, column=col).value
                if v is not None and v != "":
                    try:
                        vals.append(round(float(str(v)), 2))
                    except (ValueError, TypeError):
                        break
                else:
                    break
            if vals:
                budgets[name] = vals

    data["budgets"]            = budgets
    data["budget_race_names"]  = budget_race_names

    # ── Reference Tables sheet — overall standings positions per race ─────────
    # Anchor: find the row with "Position changes" in col A, then header is next row
    ws_ref2 = wb.worksheets[9]
    pos_header_row = None
    for r in range(1, ws_ref2.max_row + 1):
        cell_val = ws_ref2.cell(row=r, column=1).value
        if cell_val and "Position changes" in str(cell_val):
            pos_header_row = r + 1   # header row is immediately after the label
            break

    pos_race_names = []
    standings_positions = {}
    if pos_header_row:
        for col in range(2, ws_ref2.max_column + 1):
            hdr = ws_ref2.cell(row=pos_header_row, column=col).value
            if hdr and str(hdr).strip():
                pos_race_names.append(str(hdr).strip())
            else:
                break
        for r in range(pos_header_row + 1, pos_header_row + 15):
            nm = ws_ref2.cell(row=r, column=1).value
            if not nm:
                break
            nm = str(nm).strip()
            vals = []
            for col in range(2, 2 + len(pos_race_names)):
                v = ws_ref2.cell(row=r, column=col).value
                try:
                    vals.append(int(float(str(v))) if v is not None else None)
                except (ValueError, TypeError):
                    vals.append(None)
            standings_positions[nm] = vals

    data["pos_race_names"]       = pos_race_names
    data["standings_positions"]  = standings_positions

    return data


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — COMPUTE DERIVED DATA
# ─────────────────────────────────────────────────────────────────────────────

def compute(data):
    managers    = data["managers"]
    races_done  = data["races_done"]
    finish_dist = data["finish_dist"]
    chips_used  = data["chips_used"]
    budgets     = data["budgets"]
    budget_races= data["budget_race_names"]
    lineups     = data["lineups"]

    n_done = len(races_done)
    n_total = len(data["races"])

    # Sort managers by rank for consistent ordering
    managers_sorted = sorted(managers, key=lambda m: m["rank"])

    # ── Points progression (cumulative per race) ─────────────────────────────
    # data[0] = 0 (pre-season), then cumulative after each completed race
    for m in managers_sorted:
        cum = [0]
        running = 0
        for race in races_done:
            pts = m["scores"].get(race["name"])
            if pts is not None:
                running += pts
            cum.append(running)
        m["cumulative"] = cum

    # ── Per-race rankings ────────────────────────────────────────────────────
    for race in races_done:
        rname = race["name"]
        scores_this_race = [(m["name"], m["scores"].get(rname, 0)) for m in managers_sorted]
        scores_this_race.sort(key=lambda x: -x[1])
        race["ranking"] = scores_this_race  # [(name, pts), ...]
        race["winner"]  = scores_this_race[0][0] if scores_this_race else ""

    # ── Podiums — derived from per-race rankings above, same source Race Breakdown
    # uses. Replaces the Podiums worksheet read so the two pages can't drift apart
    # (e.g. when the Podiums sheet lags behind a newly-entered race's scores).
    data["podiums"] = [
        {
            "race":   race["name"],
            "first":  race["ranking"][0][0] if len(race["ranking"]) > 0 else "",
            "second": race["ranking"][1][0] if len(race["ranking"]) > 1 else "",
            "third":  race["ranking"][2][0] if len(race["ranking"]) > 2 else "",
        }
        for race in races_done
    ]

    # ── Position history — from Reference Tables (read in read_workbook) ──────
    pos_race_names      = data.get("pos_race_names", [])
    standings_positions = data.get("standings_positions", {})

    for m in managers_sorted:
        name = m["name"]
        raw_positions = standings_positions.get(name, [])
        positions = []
        for i, race in enumerate(races_done):
            rname = race["name"]
            if rname in pos_race_names:
                idx = pos_race_names.index(rname)
                if idx < len(raw_positions) and raw_positions[idx] is not None:
                    positions.append(raw_positions[idx])
                else:
                    # fallback: derive from cumulative scores
                    cum_scores = {mm["name"]: sum(mm["scores"].get(rd["name"], 0)
                                  for rd in races_done[:i+1]) for mm in managers_sorted}
                    sorted_cum = sorted(cum_scores.items(), key=lambda x: -x[1])
                    pos_map = {n: j+1 for j, (n, _) in enumerate(sorted_cum)}
                    positions.append(pos_map.get(name))
            else:
                positions.append(None)
        m["positions"] = positions

    # ── Chip usage per manager ───────────────────────────────────────────────
    chip_race_usage = data.get("chip_race_usage", {})
    for m in managers_sorted:
        name = m["name"]
        used = chips_used.get(name, {})
        m["chips"] = {chip: used.get(chip, False) for chip in CHIP_ORDER}
        used_list = [c for c in CHIP_ORDER if m["chips"].get(c)]
        m["chip_label"] = used_list[0] if len(used_list) == 1 else (
                          f"{len(used_list)} chips" if used_list else None)
        m["chip_bg"] = CHIP_STYLES[used_list[0]]["bg"] if len(used_list) == 1 else None
        m["chip_tc"] = CHIP_STYLES[used_list[0]]["tc"] if len(used_list) == 1 else None
        # Per-race chip: {race_name: chip_name} — only races where chip was actually used
        m["chip_by_race"] = chip_race_usage.get(name, {})

    # ── Budget data ──────────────────────────────────────────────────────────
    for m in managers_sorted:
        name = m["name"]
        bvals = budgets.get(name, [100.0])
        m["budgets"] = bvals
        # bvals[0] = Pre-Season start, remaining = post-race values
        m["budget_current"] = bvals[-1] if bvals else 100.0
        m["budget_start"]   = bvals[0]  if bvals else 100.0
        m["budget_vs_start"] = round(m["budget_current"] - m["budget_start"], 2) if bvals else 0
        m["budget_last_change"] = round(bvals[-1] - bvals[-2], 2) if len(bvals) >= 2 else 0

    # ── Lineups ─────────────────────────────────────────────────────────────
    for m in managers_sorted:
        m["lineups"] = lineups.get(m["name"], {})

    # ── Stats highlights ────────────────────────────────────────────────────
    all_race_scores = [(m["name"], rname, pts)
                       for m in managers_sorted
                       for rname, pts in m["scores"].items()]

    if all_race_scores:
        best  = max(all_race_scores, key=lambda x: x[2])
        worst = min(all_race_scores, key=lambda x: x[2])
    else:
        best = worst = ("—", "—", 0)

    best_avg  = max(managers_sorted, key=lambda m: m["avg"])
    worst_avg = min(managers_sorted, key=lambda m: m["avg"])

    # Most consistent = lowest std dev of per-race finish positions (not overall standings)
    # Compute per-race finish position for each manager from scores
    import statistics
    for m in managers_sorted:
        race_finishes = []
        for race in races_done:
            rname = race["name"]
            scores_this = {mm["name"]: mm["scores"].get(rname) for mm in managers_sorted
                          if mm["scores"].get(rname) is not None}
            if m["name"] in scores_this:
                sorted_scores = sorted(scores_this.items(), key=lambda x: -x[1])
                pos_map = {n: i+1 for i, (n,_) in enumerate(sorted_scores)}
                race_finishes.append(pos_map[m["name"]])
        m["race_finishes"] = race_finishes

    def finish_std(m):
        valid = m.get("race_finishes", [])
        if len(valid) < 2:
            return 99
        return statistics.stdev(valid)

    most_consistent = min(managers_sorted, key=finish_std)
    pos_desc = ", ".join(f"P{p}" for p in most_consistent["race_finishes"])

    # Biggest swing = largest overall standings position change between consecutive races
    biggest_swing_val = 0
    biggest_swing_m = managers_sorted[0]
    for m in managers_sorted:
        pos = [p for p in m["positions"] if p is not None]
        for i in range(1, len(pos)):
            swing = abs(pos[i] - pos[i-1])
            if swing > biggest_swing_val:
                biggest_swing_val = swing
                biggest_swing_m = m

    data["highlights"] = {
        "best":            best,
        "worst":           worst,
        "best_avg":        best_avg,
        "worst_avg":       worst_avg,
        "most_consistent": (most_consistent, pos_desc),
        "biggest_swing":   (biggest_swing_m, biggest_swing_val),
    }

    data["managers_sorted"] = managers_sorted
    data["n_done"]  = n_done
    data["n_total"] = n_total
    data["races_done"] = races_done

    return data


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — GENERATE HTML PANELS
# ─────────────────────────────────────────────────────────────────────────────

def js(v):
    """Convert Python value to safe JS literal."""
    if v is None:     return "null"
    if isinstance(v, bool): return "true" if v else "false"
    if isinstance(v, str):  return "'" + v.replace("'", "\\'") + "'"
    if isinstance(v, list): return "[" + ",".join(js(x) for x in v) + "]"
    if isinstance(v, dict): return "{" + ",".join(f'"{k}":{js(vv)}' for k, vv in v.items()) + "}"
    return str(v)

def chip_pill(label, bg, tc, small=False):
    size = "9px" if small else "10px"
    pad  = "1px 5px" if small else "1px 7px"
    return (f'<span class="chip-pill" style="background:{bg};color:{tc};'
            f'font-size:{size};padding:{pad}">{label}</span>')


# ── 01 Leaderboard ───────────────────────────────────────────────────────────
def panel_leaderboard(data):
    M   = data["managers_sorted"]
    RD  = data["races_done"]
    NT  = data["n_total"]
    ND  = data["n_done"]

    leader = M[0]["total"] if M else 0
    last   = M[-1]["total"] if M else 0
    gap_span = leader - last

    # Latest podiums line
    medal_pill_styles = [
        "background:#2a2200;border:1px solid #FFD700;color:#FFD700",
        "background:#1e1e1e;border:1px solid #C0C0C0;color:#C0C0C0",
        "background:#221a12;border:1px solid #CD7F32;color:#CD7F32",
    ]
    medal_labels = ["1st", "2nd", "3rd"]
    pod_cards = ""
    for pod in reversed([p for p in data["podiums"] if p["first"] or p["second"] or p["third"]][-3:]):
        names = [pod["first"], pod["second"], pod["third"]]
        if any(names):
            slots_html = ""
            for idx, name in enumerate(names):
                if name:
                    mgr_color = MANAGER_COLOURS.get(name, "#888")
                    slots_html += (
                        f'<span style="display:inline-flex;align-items:center;gap:5px;'
                        f'{medal_pill_styles[idx]};border-radius:5px;padding:2px 7px;font-size:11px;font-weight:500;white-space:nowrap">'
                        f'<span style="font-size:10px;color:inherit;opacity:0.75">{medal_labels[idx]}</span>'
                        f'<span style="color:{mgr_color}">{name}</span>'
                        f'</span>'
                    )
            pod_cards += (
                f'<div class="podium-card">'
                f'<div class="podium-pos">{pod["race"]}</div>'
                f'<div style="display:flex;gap:4px;flex-wrap:wrap;margin-top:5px">{slots_html}</div>'
                f'</div>'
            )

    # Standings rows
    medal_styles = [
        "background:#FFD700;color:#7a5800",
        "background:#C0C0C0;color:#4a4a4a",
        "background:#CD7F32;color:#5a2d00",
    ]
    rows_html = ""
    for i, m in enumerate(M):
        pct   = round(m["total"] / leader * 100, 1) if leader else 0
        medal = medal_styles[i] if i < 3 else "background:#2a2a2a;color:#888"
        gap   = "Leader" if i == 0 else str(M[i]["total"] - M[0]["total"])
        rows_html += f"""<div class="row">
  <div class="pos-badge" style="{medal}">{i+1}</div>
  <div><div class="manager-name">{m['name']}</div><div class="team-name">{m['team']}</div>
  <div class="bar-bg"><div class="bar-fill" style="width:{pct}%;background:{m['color']}"></div></div></div>
  <div class="gap-col">{gap}</div>
  <div class="pts-col">{m['total']}</div>
</div>"""

    # Progress chart
    labels  = ["Pre-season"] + [r["name"] for r in RD]
    datasets = []
    for i, m in enumerate(M):
        dash = DASH_PATTERNS[i % len(DASH_PATTERNS)]
        datasets.append({
            "label": m["name"], "data": m["cumulative"],
            "borderColor": m["color"], "borderDash": dash,
            "borderWidth": 2, "pointBackgroundColor": m["color"],
            "pointRadius": 4, "pointHoverRadius": 6,
            "fill": False, "tension": 0.2
        })

    legend_html = "".join(
        f'<span><span style="width:10px;height:10px;border-radius:2px;background:{m["color"]};display:inline-block"></span>{m["name"]}</span>'
        for m in M)

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">F1 Fantasy League — Season Leaderboard · {ND} of {NT} races complete</div>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">{ND}</div><div class="mc-lbl">Races complete</div></div>
  <div class="mc"><div class="mc-val">{NT - ND}</div><div class="mc-lbl">Races remaining</div></div>
  <div class="mc"><div class="mc-val">{leader}</div><div class="mc-lbl">Pts — leader</div></div>
  <div class="mc"><div class="mc-val">{gap_span}</div><div class="mc-lbl">Pts gap P1 to P{len(M)}</div></div>
</div>
<div class="section-label">Latest podiums</div>
<div class="podium-row">{pod_cards}</div>
<div class="section-label">Standings</div>
<div class="card">{rows_html}</div>
<div class="hint">Bar shows points as % of leader's total ({leader} pts)</div>
<div class="section-label">Points progression</div>
<div style="position:relative;height:280px"><canvas id="progressChart"></canvas></div>
<div class="legend" id="legend-leaderboard"></div>
<script>
(function(){{
const datasets={js(datasets)};
new Chart(document.getElementById('progressChart'),{{
  type:'line',
  data:{{labels:{js(labels)},datasets:datasets}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y}} pts`}}}}}},
    scales:{{x:{{ticks:{{color:'#888'}},grid:{{color:'rgba(255,255,255,0.06)'}}}},
             y:{{ticks:{{color:'#888'}},grid:{{color:'rgba(255,255,255,0.06)'}}}}}}}}
}});
const leg=document.getElementById('legend-leaderboard');
leg.innerHTML={js(legend_html)};
}})();
</script>"""


# ── 02 Race Breakdown ─────────────────────────────────────────────────────────
def panel_race_breakdown(data):
    M   = data["managers_sorted"]
    RD  = data["races_done"]
    ND  = data["n_done"]
    NT  = data["n_total"]

    all_scores = [pts for m in M for pts in m["scores"].values()]
    high = max(all_scores) if all_scores else 0
    low  = min(all_scores) if all_scores else 0
    avg  = round(sum(all_scores) / len(all_scores)) if all_scores else 0

    high_m = next((f"{m['name']}, R{i+1}" for i, race in enumerate(RD)
                   for m in M if m["scores"].get(race["name"]) == high), "")

    medal_styles = ["background:#FFD700;color:#7a5800",
                    "background:#C0C0C0;color:#4a4a4a",
                    "background:#CD7F32;color:#5a2d00"]
    slot_styles  = ["background:#2a2200;border:0.5px solid #FFD700",
                    "background:#222;border:0.5px solid #C0C0C0",
                    "background:#221a12;border:0.5px solid #CD7F32"]
    ordinals     = ["1st","2nd","3rd","4th","5th","6th","7th","8th","9th"]

    race_cards_html = ""
    for race in RD:
        rname = race["name"]
        rnd   = race["round"]
        scores_here = [(m, m["scores"].get(rname, 0)) for m in M if rname in m["scores"]]
        scores_here.sort(key=lambda x: -x[1])
        max_pts = scores_here[0][1] if scores_here else 1

        podium_html = ""
        for j in range(min(3, len(scores_here))):
            mm, pts = scores_here[j]
            podium_html += (f'<div class="slot" style="{slot_styles[j]}">'
                            f'<div class="medal" style="{medal_styles[j]}">{j+1}</div>'
                            f'<div><div class="slot-name" style="color:{mm["color"]}">{mm["name"]}</div>'
                            f'<div class="slot-pts">{pts} pts</div></div></div>')

        score_rows = ""
        for j, (mm, pts) in enumerate(scores_here):
            pct = round(max(0, pts) / max_pts * 100) if max_pts > 0 else 0
            bg  = ["#FFD700","#C0C0C0","#CD7F32"][j] if j < 3 else "#2a2a2a"
            tc  = ["#7a5800","#4a4a4a","#5a2d00"][j] if j < 3 else "#888"
            race_chip = mm["chip_by_race"].get(rname)
            if race_chip and race_chip in CHIP_STYLES:
                pill = chip_pill(race_chip, CHIP_STYLES[race_chip]["bg"], CHIP_STYLES[race_chip]["tc"], small=True)
            else:
                pill = ""
            score_rows += (f'<div class="score-row">'
                           f'<div class="pos-dot" style="background:{bg};color:{tc}">{j+1}</div>'
                           f'<div><div class="score-name">{mm["name"]}{pill}</div>'
                           f'<div class="bar-bg"><div class="bar-fill" style="width:{pct}%;background:{mm["color"]}"></div></div></div>'
                           f'<div class="score-pts">{pts}</div></div>')

        all_pts   = [s for _, s in scores_here]
        rng       = max(all_pts) - min(all_pts) if all_pts else 0
        r_avg     = round(sum(all_pts) / len(all_pts)) if all_pts else 0
        negatives = sum(1 for s in all_pts if s < 0)
        neg_note  = f"{negatives} went negative" if negatives else "Everyone positive"

        race_cards_html += f"""<div class="race-card">
  <div class="race-header"><span class="race-title">{rname}</span><span class="race-round">Round {rnd}</span></div>
  <div class="podium-strip">{podium_html}</div>
  {score_rows}
  <div><span class="tag">Range: {rng} pts</span><span class="tag">Avg: {r_avg} pts</span><span class="tag">{neg_note}</span></div>
</div>"""

    # Grouped bar chart data
    bar_datasets = []
    for m in M:
        race_pts = [m["scores"].get(r["name"], 0) for r in RD]
        bar_datasets.append({"label": m["name"], "data": race_pts,
                              "backgroundColor": m["color"], "borderWidth": 0})
    race_labels = [r["name"] for r in RD]
    legend_html = "".join(
        f'<span><span style="width:10px;height:10px;border-radius:2px;background:{m["color"]};display:inline-block"></span>{m["name"]}</span>'
        for m in M)

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Race Breakdown · {ND} of {NT} races complete</div>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">{ND}</div><div class="mc-lbl">Races complete</div></div>
  <div class="mc"><div class="mc-val">{high}</div><div class="mc-lbl">Highest single race ({high_m})</div></div>
  <div class="mc"><div class="mc-val">{low}</div><div class="mc-lbl">Lowest single race</div></div>
  <div class="mc"><div class="mc-val">{avg}</div><div class="mc-lbl">Avg points per race</div></div>
</div>
<div class="section-label">Round by round</div>
{race_cards_html}
<div class="section-label">Points per race — all managers</div>
<div style="position:relative;height:300px">
  <div id="barChartYAxis" style="position:absolute;top:0;left:0;width:48px;height:100%;z-index:2;background:#0f0f0f;pointer-events:none"></div>
  <div id="barChartScroll" style="overflow-x:auto;-webkit-overflow-scrolling:touch;height:100%;padding-left:48px;box-sizing:border-box">
    <div style="position:relative;height:100%;min-width:{max(len(race_labels)*90, 400)}px">
      <canvas id="barChart"></canvas>
    </div>
  </div>
</div>
<div class="legend" id="legend-race"></div>
<script>
(function(){{
const nRaces={len(race_labels)};
const chart=new Chart(document.getElementById('barChart'),{{
  type:'bar',
  data:{{labels:{js(race_labels)},datasets:{js(bar_datasets)}}},
  options:{{responsive:true,maintainAspectRatio:false,
    layout:{{padding:{{top:10,left:0}}}},
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y}} pts`}}}}}},
    scales:{{
      x:{{ticks:{{color:'#888',maxRotation:30}},grid:{{color:'rgba(255,255,255,0.06)'}}}},
      y:{{ticks:{{color:'#888',display:false}},grid:{{color:'rgba(255,255,255,0.06)'}},border:{{display:false}}}}
    }}}}
}});
function freezeYAxis(){{
  const yAxis=document.getElementById('barChartYAxis');
  const cvs=document.getElementById('barChart');
  const scale=chart.scales.y;
  if(!scale)return;
  const dpr=window.devicePixelRatio||1;
  const cssH=cvs.clientHeight||300;
  const fc=document.createElement('canvas');
  fc.width=48*dpr; fc.height=cssH*dpr;
  fc.style.width='48px'; fc.style.height=cssH+'px';
  const ctx2=fc.getContext('2d');
  ctx2.scale(dpr,dpr);
  ctx2.fillStyle='#0f0f0f';
  ctx2.fillRect(0,0,48,cssH);
  const ticks=scale.ticks;
  const top=scale.top; const bottom=scale.bottom;
  const range=bottom-top;
  const vMin=scale.min; const vMax=scale.max;
  ticks.forEach(t=>{{
    const v=t.value;
    const yPx=top+((vMax-v)/(vMax-vMin))*range;
    ctx2.fillStyle='#888';
    ctx2.font='11px -apple-system,BlinkMacSystemFont,sans-serif';
    ctx2.textAlign='right';
    ctx2.fillText(t.label,44,yPx+4);
  }});
  yAxis.innerHTML='';
  yAxis.appendChild(fc);
}}
chart.options.animation={{onComplete:freezeYAxis}};
chart.update();
document.getElementById('legend-race').innerHTML={js(legend_html)};
}})();
</script>"""


# ── 03 Head-to-Head ───────────────────────────────────────────────────────────
def panel_h2h(data):
    M  = data["managers_sorted"]
    RD = data["races_done"]

    manager_data_js = {}
    for m in M:
        manager_data_js[m["name"]] = {
            "team":      m["team"],
            "color":     m["color"],
            "pos":       m["rank"],
            "total":     m["total"],
            "races":     [m["scores"].get(r["name"], 0) for r in RD],
            "podiums":   [1 if any(
                             pod["race"] == r["name"] and
                             pod[pos_key] == m["name"]
                             for pod in data["podiums"]
                             for pos_key in ["first","second","third"]
                           ) else 0
                          for r in RD],
            "wins":      [1 if any(
                             pod["race"] == r["name"] and pod["first"] == m["name"]
                             for pod in data["podiums"]
                           ) else 0
                          for r in RD],
            "chips":     {c: m["chips"].get(c, False) for c in CHIP_ORDER},
            "chipLabel":   m["chip_label"],
            "chipBg":      m["chip_bg"],
            "chipTc":      m["chip_tc"],
            "chipByRace":  m["chip_by_race"],   # {race_name: chip_name}
        }

    names_js = js(list(manager_data_js.keys()))
    data_js  = js(manager_data_js)
    chips_js = js([{"key": c, "label": c,
                    "bg": CHIP_STYLES[c]["bg"], "tc": CHIP_STYLES[c]["tc"],
                    "desc": {"Limitless":"Unlimited transfers","Wildcard":"Unlimited changes",
                             "Final Fix":"Post-quali changes","Auto Pilot":"Auto DRS boost",
                             "No Negative":"No negative pts","Extra DRS":"3x pts booster"}[c]}
                   for c in CHIP_ORDER])
    race_names_js = js([r["name"] for r in RD])

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Head-to-Head Comparison</div>
<div style="display:flex;gap:12px;align-items:center;margin-bottom:1.5rem;flex-wrap:wrap">
  <div style="display:flex;align-items:center;gap:8px"><label>Manager A</label><select id="selA" onchange="ucH2HSync('A')"></select></div>
  <div style="font-size:13px;color:#888;font-weight:500">vs</div>
  <div style="display:flex;align-items:center;gap:8px"><label>Manager B</label><select id="selB" onchange="ucH2HSync('B')"></select></div>
</div>
<div id="vs-header" class="vs-header"></div>
<div class="section-label" style="margin-top:0">Season stats</div>
<div class="stat-grid" id="stat-grid"></div>
<div class="section-label">Chip usage</div>
<div class="card" id="chip-table"></div>
<div class="section-label">Race by race</div>
<div class="card" id="race-rows"></div>
<div class="section-label">Points progression</div>
<div style="position:relative;height:220px"><canvas id="h2hChart"></canvas></div>
<script>
(function(){{
const chips={chips_js};
const chipStyles={js({c: {"bg": CHIP_STYLES[c]["bg"], "tc": CHIP_STYLES[c]["tc"]} for c in CHIP_ORDER})};
const mgrs={data_js};
const raceNames={race_names_js};
const names={names_js};
let chart=null;
const selA=document.getElementById('selA'),selB=document.getElementById('selB');
function populateSelect(sel,exclude,cur){{
  const prev=cur||sel.value;sel.innerHTML='';
  names.filter(n=>n!==exclude).forEach(n=>{{const o=document.createElement('option');o.value=n;o.textContent=n;if(n===prev)o.selected=true;sel.appendChild(o);}});
  if(!sel.value)sel.value=sel.options[0].value;
}}
function syncSelects(changed){{if(changed==='A')populateSelect(selB,selA.value,selB.value);else populateSelect(selA,selB.value,selA.value);render();}}
window.ucH2HSync=syncSelects;
populateSelect(selA,null,names[0]);
if(names.length>1)populateSelect(selB,names[0],names[1]);
else populateSelect(selB,null,names[0]);
function tickSvg(used,color){{
  if(used)return`<div class="tick" style="background:${{color}}22"><svg width="11" height="11" viewBox="0 0 12 12" fill="none"><path d="M2 6l3 3 5-5" stroke="${{color}}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg></div>`;
  return`<div class="tick" style="background:#1e1e1e"><svg width="11" height="11" viewBox="0 0 12 12" fill="none"><path d="M4 4l4 4M8 4l-4 4" stroke="#555" stroke-width="1.5" stroke-linecap="round"/></svg></div>`;
}}
function raceChipPill(mgr, raceName){{
  const label=mgr.chipByRace&&mgr.chipByRace[raceName];
  if(!label)return'';
  const st=chipStyles[label]||{{bg:'#222',tc:'#aaa'}};
  return`<span style="display:inline-block;font-size:9px;padding:1px 5px;border-radius:20px;font-weight:500;margin-left:4px;background:${{st.bg}};color:${{st.tc}}">${{label}}</span>`;
}}
function render(){{
  const nA=selA.value,nB=selB.value,A=mgrs[nA],B=mgrs[nB];
  const aRW=A.wins.reduce((s,v)=>s+v,0),bRW=B.wins.reduce((s,v)=>s+v,0);
  const aP=A.podiums.reduce((s,v)=>s+v,0),bP=B.podiums.reduce((s,v)=>s+v,0);
  const aL=A.total>=B.total;
  document.getElementById('vs-header').innerHTML=`
    <div class="vs-card${{aL?' winner':''}}" style="${{aL?`border-color:${{A.color}}`:''}}">
      <div class="vs-name" style="color:${{A.color}}">${{nA}}</div><div class="vs-team">${{A.team}}</div>
      <div class="vs-pts" style="color:${{A.color}}">${{A.total}}</div><div class="vs-pos">P${{A.pos}} overall</div></div>
    <div class="vs-divider">vs</div>
    <div class="vs-card${{!aL?' winner':''}}" style="${{!aL?`border-color:${{B.color}}`:''}}">
      <div class="vs-name" style="color:${{B.color}}">${{nB}}</div><div class="vs-team">${{B.team}}</div>
      <div class="vs-pts" style="color:${{B.color}}">${{B.total}}</div><div class="vs-pos">P${{B.pos}} overall</div></div>`;
  const rows=[[A.total,'Total pts',B.total],[Math.round(A.total/Math.max(1,raceNames.length)),'Avg per race',Math.round(B.total/Math.max(1,raceNames.length))],[Math.max(...A.races),'Best race',Math.max(...B.races)],[Math.min(...A.races),'Worst race',Math.min(...B.races)],[aRW,'Race wins',bRW],[aP,'Podiums',bP]];
  document.getElementById('stat-grid').innerHTML=rows.map(([av,lbl,bv])=>{{
    const aH=av>bv,bH=bv>av;
    return`<div class="stat-val${{aH?' hi':''}}" style="${{aH?`border-color:${{A.color}}`:''}}">  ${{av}}</div><div class="stat-label">${{lbl}}</div><div class="stat-val${{bH?' hi':''}}" style="${{bH?`border-color:${{B.color}}`:''}}">${{bv}}</div>`;
  }}).join('');
  const aU=chips.filter(c=>A.chips[c.key]).length,bU=chips.filter(c=>B.chips[c.key]).length;
  document.getElementById('chip-table').innerHTML=
    `<div class="chip-row" style="padding:10px 0"><div style="font-size:11px;font-weight:500;color:${{A.color}};text-align:center">${{nA}}<br><span style="font-size:10px;color:#888;font-weight:400">${{aU}} used</span></div><div style="font-size:11px;color:#888;text-align:center">Chip</div><div style="font-size:11px;font-weight:500;color:${{B.color}};text-align:center">${{nB}}<br><span style="font-size:10px;color:#888;font-weight:400">${{bU}} used</span></div></div>`+
    chips.map(c=>`<div class="chip-row"><div style="display:flex;justify-content:center">${{tickSvg(A.chips[c.key],A.color)}}</div><div style="text-align:center"><span class="chip-pill" style="background:${{c.bg}};color:${{c.tc}}">${{c.label}}</span><div style="font-size:10px;color:#555;margin-top:3px">${{c.desc}}</div></div><div style="display:flex;justify-content:center">${{tickSvg(B.chips[c.key],B.color)}}</div></div>`).join('');
  const maxP=Math.max(...[...A.races,...B.races].map(Math.abs),1);
  function chipPill(m){{
    if(!m.chipLabel)return'';
    return`<span style="display:inline-block;font-size:9px;padding:1px 5px;border-radius:20px;font-weight:500;margin-left:4px;background:${{m.chipBg}};color:${{m.chipTc}}">${{m.chipLabel}}</span>`;
  }}
  document.getElementById('race-rows').innerHTML=raceNames.map((race,i)=>{{
    const ap=A.races[i],bp=B.races[i],aW=ap>bp;
    const aPct=Math.round(Math.abs(ap)/maxP*100),bPct=Math.round(Math.abs(bp)/maxP*100);
    return`<div class="race-row">
      <div style="text-align:right"><div style="font-size:14px;font-weight:500;color:${{aW?A.color:'#888'}}">${{ap}}</div><div style="font-size:10px;margin-top:2px;text-align:right">${{raceChipPill(A,race)}}</div><div style="display:flex;justify-content:flex-end;margin-top:4px"><div style="height:6px;border-radius:3px;width:${{aPct}}%;background:${{A.color}};opacity:${{aW?1:0.4}}"></div></div></div>
      <div class="race-label">${{race}}</div>
      <div><div style="font-size:14px;font-weight:500;color:${{!aW?B.color:'#888'}}">${{bp}}</div><div style="font-size:10px;margin-top:2px">${{raceChipPill(B,race)}}</div><div style="display:flex;margin-top:4px"><div style="height:6px;border-radius:3px;width:${{bPct}}%;background:${{B.color}};opacity:${{!aW?1:0.4}}"></div></div></div></div>`;
  }}).join('');
  const cumA=A.races.reduce((acc,v)=>[...acc,acc[acc.length-1]+v],[0]);
  const cumB=B.races.reduce((acc,v)=>[...acc,acc[acc.length-1]+v],[0]);
  if(chart)chart.destroy();
  chart=new Chart(document.getElementById('h2hChart'),{{type:'line',data:{{labels:['Pre-season',...raceNames],datasets:[
    {{label:nA,data:cumA,borderColor:A.color,backgroundColor:A.color+'22',borderWidth:2.5,pointBackgroundColor:A.color,pointRadius:5,fill:true,tension:0.2}},
    {{label:nB,data:cumB,borderColor:B.color,backgroundColor:B.color+'22',borderWidth:2.5,pointBackgroundColor:B.color,pointRadius:5,fill:true,tension:0.2}}
  ]}},options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y}} pts`}}}}}},
    scales:{{x:{{ticks:{{color:'#888'}},grid:{{color:'rgba(255,255,255,0.06)'}}}},y:{{ticks:{{color:'#888'}},grid:{{color:'rgba(255,255,255,0.06)'}}}}}}}}
  }});
}}
render();
}})();
</script>"""


# ── 04 Budget Tracker ─────────────────────────────────────────────────────────
def panel_budget(data):
    M      = data["managers_sorted"]
    b_races= data["budget_race_names"]
    RD     = data["races_done"]

    # Sort by current budget desc
    sorted_m = sorted(M, key=lambda m: -m["budget_current"])
    max_b    = max(m["budget_current"] for m in sorted_m) if sorted_m else 100
    min_b    = min(m["budget_current"] for m in sorted_m) if sorted_m else 100
    avg_b    = round(sum(m["budget_current"] for m in M) / len(M), 1) if M else 100

    rows_html = ""
    for m in sorted_m:
        cur  = m["budget_current"]
        vs   = round(cur - m["budgets"][0], 2) if m["budgets"] else 0
        lc   = m.get("budget_last_change", 0)
        pct  = round(cur / max_b * 100, 1)
        vs_c = "pos" if vs > 0 else ("neg" if vs < 0 else "neu")
        lc_c = "pos" if lc > 0 else ("neg" if lc < 0 else "neu")
        vs_s = "+" if vs > 0 else ""
        lc_s = "+" if lc > 0 else ""
        rows_html += f"""<div class="brow">
  <div style="display:flex;align-items:center;justify-content:center"><div class="dot" style="background:{m['color']}"></div></div>
  <div><div class="bname">{m['name']}</div><div class="bar-bg"><div class="bar-fill" style="width:{pct}%;background:{m['color']}"></div></div></div>
  <div class="bcol" style="font-weight:500">{cur:.1f}m</div>
  <div class="bcol {vs_c}">{vs_s}{vs:.1f}m</div>
  <div class="bcol {lc_c}">{lc_s}{lc:.1f}m</div>
</div>"""

    # Budget timeline chart
    timeline_labels = b_races[:max(len(m["budgets"]) for m in M)] if M else []
    timeline_datasets = []
    for i, m in enumerate(M):
        dash = DASH_PATTERNS[i % len(DASH_PATTERNS)]
        timeline_datasets.append({
            "label": m["name"], "data": m["budgets"],
            "borderColor": m["color"], "borderDash": dash,
            "borderWidth": 2, "pointBackgroundColor": m["color"],
            "pointRadius": 4, "pointHoverRadius": 6,
            "fill": False, "tension": 0.2
        })

    # Budget change per race — HTML table (no canvas, no lifecycle issues)
    max_races = max((len(m["budgets"]) for m in M), default=1)
    # Only show races that have actually completed (budget data exists for all managers)
    # change index i means: budgets[i] - budgets[i-1], so race label is budget_race_names[i]
    n_change_cols = max_races - 1  # number of completed race changes
    change_race_names = b_races[1:max_races] if len(b_races) >= max_races else b_races[1:]

    # Build per-manager change lists (only completed races)
    manager_changes = []
    for m in M:
        changes = [round(m["budgets"][i] - m["budgets"][i-1], 2)
                   for i in range(1, len(m["budgets"]))]
        manager_changes.append((m, changes))

    # Build table header
    race_th = "".join(
        f'<th style="font-size:10px;color:#555;font-weight:500;text-align:right;padding:4px 8px;white-space:nowrap">{rn}</th>'
        for rn in change_race_names)

    # Build table rows
    table_rows = ""
    for m, changes in manager_changes:
        cells = ""
        for c in changes:
            if c > 0:
                color = "#1D9E75"; sign = "+"
            elif c < 0:
                color = "#E24B4A"; sign = ""
            else:
                color = "#555"; sign = ""
            cells += f'<td style="font-size:12px;font-weight:500;color:{color};text-align:right;padding:5px 8px;white-space:nowrap">{sign}{c:.1f}m</td>'
        table_rows += f"""<tr style="border-bottom:0.5px solid #2a2a2a">
  <td style="padding:5px 8px;white-space:nowrap;position:sticky;left:0;background:#1a1a1a;z-index:1">
    <div style="display:flex;align-items:center;gap:7px">
      <div style="width:8px;height:8px;border-radius:50%;background:{m['color']};flex-shrink:0"></div>
      <span style="font-size:13px;font-weight:500">{m['name']}</span>
    </div>
  </td>
  {cells}
</tr>"""

    budget_change_html = f"""<div class="card" style="padding:4px 0;overflow-x:auto;-webkit-overflow-scrolling:touch">
  <table style="border-collapse:collapse;min-width:max(100%,{max(len(change_race_names)*90+120, 300)}px)">
    <thead>
      <tr style="border-bottom:0.5px solid #2a2a2a">
        <th style="font-size:10px;color:#555;font-weight:500;text-align:left;padding:4px 8px;position:sticky;left:0;background:#1a1a1a;z-index:1;min-width:90px">Manager</th>
        {race_th}
      </tr>
    </thead>
    <tbody>{table_rows}</tbody>
  </table>
</div>"""

    legend_html = "".join(
        f'<span><span style="width:10px;height:10px;border-radius:2px;background:{m["color"]};display:inline-block"></span>{m["name"]}</span>'
        for m in M)

    import math
    y_max = math.ceil((max_b + 2) / 2) * 2   # next even number at least 2m above highest budget
    y_min = math.floor((min_b - 2) / 2) * 2  # next even number at least 2m below lowest budget
    y_min = max(80, y_min)   # never go below 80m
    y_max = min(120, y_max)  # never go above 120m

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Budget Tracker · Team values across the season</div>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">{max_b:.1f}m</div><div class="mc-lbl">Highest budget ({sorted_m[0]['name']})</div></div>
  <div class="mc"><div class="mc-val">{min_b:.1f}m</div><div class="mc-lbl">Lowest budget ({sorted_m[-1]['name']})</div></div>
  <div class="mc"><div class="mc-val">{round(max_b - min_b, 1)}m</div><div class="mc-lbl">Largest gap</div></div>
  <div class="mc"><div class="mc-val">{avg_b}m</div><div class="mc-lbl">League average</div></div>
</div>
<div class="section-label">Current budgets</div>
<div class="card">
  <div style="display:grid;grid-template-columns:26px 1fr 52px 60px 60px;gap:10px;padding:6px 0 2px">
    <div></div><div></div>
    <div class="col-hdr">Current</div><div class="col-hdr">vs start</div><div class="col-hdr">Last race</div>
  </div>
  {rows_html}
</div>
<div class="section-label">Budget timeline</div>
<div style="position:relative;height:280px"><canvas id="budgetChart"></canvas></div>
<div class="legend" id="legend-budget"></div>
<div class="section-label">Budget change per race</div>
{budget_change_html}
<script>
(function(){{
new Chart(document.getElementById('budgetChart'),{{
  type:'line',
  data:{{labels:{js(timeline_labels)},datasets:{js(timeline_datasets)}}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y.toFixed(1)}}m`}}}}}},
    scales:{{x:{{ticks:{{color:'#888'}},grid:{{color:'rgba(255,255,255,0.06)'}}}},
             y:{{min:{y_min},max:{y_max},ticks:{{color:'#888',callback:v=>v.toFixed(0)+'m'}},grid:{{color:'rgba(255,255,255,0.06)'}}}}}}}}
}});
document.getElementById('legend-budget').innerHTML={js(legend_html)};
}})();
</script>"""


# ── 05 Podiums ────────────────────────────────────────────────────────────────
def panel_podiums(data):
    M   = data["managers_sorted"]
    POD = data["podiums"]
    ND  = data["n_done"]
    NT  = data["n_total"]

    managers_with_pod = sum(1 for m in M
                            if any(p["first"]==m["name"] or p["second"]==m["name"]
                                   or p["third"]==m["name"] for p in POD))
    diff_winners = len(set(p["first"] for p in POD if p["first"]))

    medal_styles = ["background:#FFD700;color:#7a5800",
                    "background:#C0C0C0;color:#4a4a4a",
                    "background:#CD7F32;color:#5a2d00"]
    slot_styles  = ["background:#2a2200;border:0.5px solid #FFD700",
                    "background:#222;border:0.5px solid #C0C0C0",
                    "background:#221a12;border:0.5px solid #CD7F32"]

    def get_color(name):
        return MANAGER_COLOURS.get(name, "#888")

    def get_chip_pill(name, race_name):
        m = next((mm for mm in M if mm["name"] == name), None)
        if m:
            chip = m["chip_by_race"].get(race_name)
            if chip and chip in CHIP_STYLES:
                return chip_pill(chip, CHIP_STYLES[chip]["bg"], CHIP_STYLES[chip]["tc"], small=True)
        return ""

    cards_html = ""
    for pod in POD:
        filled = pod["first"] or pod["second"] or pod["third"]
        style = "" if filled else 'style="border-style:dashed;opacity:0.5"'
        slots = ""
        for j, (key, pts_key) in enumerate([("first",""), ("second",""), ("third","")]):
            name = pod[key]
            if name:
                pts = next((m["scores"].get(pod["race"], "") for m in M if m["name"] == name), "")
                pts_html = f'<div class="slot-pts">{pts} pts</div>' if pts != "" else ""
                slots += (f'<div class="slot" style="{slot_styles[j]}">'
                          f'<div class="medal" style="{medal_styles[j]}">{j+1}</div>'
                          f'<div><div class="slot-name" style="color:{get_color(name)}">{name}{get_chip_pill(name, pod["race"])}</div>{pts_html}</div></div>')
            else:
                slots += (f'<div class="slot" style="{slot_styles[j]}">'
                          f'<div class="medal" style="{medal_styles[j]}">{j+1}</div>'
                          f'<div class="tbd">TBC</div></div>')
        cards_html += f"""<div class="race-podium" {style}>
  <div class="race-header"><span class="race-title">{pod['race']}</span></div>
  <div class="podium-slots">{slots}</div>
</div>"""

    # Podium share chart
    p1 = [sum(1 for p in POD if p["first"]==m["name"]) for m in M]
    p2 = [sum(1 for p in POD if p["second"]==m["name"]) for m in M]
    p3 = [sum(1 for p in POD if p["third"]==m["name"]) for m in M]
    names_labels = [m["name"] for m in M]
    max_y = max(max(p1+p2+p3, default=0) + 1, 3)

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Podiums · Race results &amp; podium share</div>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">{ND}</div><div class="mc-lbl">Races complete</div></div>
  <div class="mc"><div class="mc-val">{managers_with_pod}</div><div class="mc-lbl">Managers with a podium</div></div>
  <div class="mc"><div class="mc-val">{diff_winners}</div><div class="mc-lbl">Different race winners</div></div>
  <div class="mc"><div class="mc-val">{len(M) - managers_with_pod}</div><div class="mc-lbl">Managers without a podium</div></div>
</div>
<div class="section-label">Podium share</div>
<div style="position:relative;height:260px"><canvas id="podiumChart"></canvas></div>
<div class="legend" id="legend-podiums"></div>
<script>
(function(){{
new Chart(document.getElementById('podiumChart'),{{
  type:'bar',
  data:{{labels:{js(names_labels)},datasets:[
    {{label:'1st',data:{js(p1)},backgroundColor:'#FFD700',borderWidth:0}},
    {{label:'2nd',data:{js(p2)},backgroundColor:'#C0C0C0',borderWidth:0}},
    {{label:'3rd',data:{js(p3)},backgroundColor:'#CD7F32',borderWidth:0}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y}}`}}}}}},
    scales:{{x:{{stacked:true,ticks:{{color:'#888'}},grid:{{display:false}}}},
             y:{{stacked:true,max:{max_y},ticks:{{color:'#888',stepSize:1}},grid:{{color:'rgba(255,255,255,0.06)'}}}}}}}}
}});
document.getElementById('legend-podiums').innerHTML=`
  <span><span style="width:10px;height:10px;border-radius:2px;background:#FFD700;display:inline-block"></span>1st place</span>
  <span><span style="width:10px;height:10px;border-radius:2px;background:#C0C0C0;display:inline-block"></span>2nd place</span>
  <span><span style="width:10px;height:10px;border-radius:2px;background:#CD7F32;display:inline-block"></span>3rd place</span>`;
}})();
</script>
<div class="section-label">Race results</div>
{cards_html}"""


# ── 06 Stats ─────────────────────────────────────────────────────────────────
def panel_stats(data):
    M   = data["managers_sorted"]
    ND  = data["n_done"]
    HL  = data["highlights"]

    total_race_scores = len(data["races_done"])
    chips_used_count  = sum(1 for m in M for c in CHIP_ORDER if m["chips"].get(c))

    # Finish distribution table
    N_FIN = data.get("n_finish_cols", len(M))

    pos_headers = "".join(
        f'<div style="font-size:9px;color:#555;text-align:center">P{i}</div>'
        for i in range(1, N_FIN + 1))
    finish_rows_html = ""
    for m in M:
        fd = data["finish_dist"].get(m["name"], [0]*N_FIN)
        cells = ""
        for i, v in enumerate(fd[:N_FIN]):
            if v == 0:
                cls = "f0"
            elif i == 0: cls = "f1"
            elif i == 1: cls = "f2"
            elif i == 2: cls = "f3"
            else:        cls = "fn"
            cells += f'<div style="display:flex;justify-content:center"><div class="fin {cls}">{v if v else ""}</div></div>'
        finish_rows_html += f"""<div class="srow">
  <div style="display:flex;align-items:center;justify-content:center"><div class="dot" style="background:{m['color']}"></div></div>
  <div class="sname">{m['name']}</div>
  {cells}<div class="total-pts">{m['total']}</div>
</div>"""

    # Chip usage rows
    chip_rows_html = ""
    for chip in CHIP_ORDER:
        cs    = CHIP_STYLES[chip]
        users = [m["name"] for m in M if m["chips"].get(chip)]
        users_html = "".join(
            f'<span style="font-size:12px;font-weight:500;color:{MANAGER_COLOURS.get(u,"#888")}">{u}</span>'
            for u in users) or '<span style="font-size:12px;color:#555">Not yet used</span>'
        chip_rows_html += (
            f'<div style="display:flex;align-items:center;gap:10px;padding:8px 0;border-bottom:0.5px solid #2a2a2a">'
            f'<span class="chip-pill" style="background:{cs["bg"]};color:{cs["tc"]};font-size:11px;padding:2px 10px">{chip}</span>'
            f'<div style="flex:1;display:flex;gap:6px;flex-wrap:wrap">{users_html}</div>'
            f'<span style="font-size:11px;color:#888">{len(users)}/{len(M)} used</span></div>')

    # Highlights
    best_m, best_r, best_pts      = HL["best"]
    worst_m, worst_r, worst_pts   = HL["worst"]
    best_avg_m                    = HL["best_avg"]
    worst_avg_m                   = HL["worst_avg"]
    cons_m, cons_pos              = HL["most_consistent"]
    swing_m, swing_val            = HL["biggest_swing"]

    def hl_row(name, desc):
        color = MANAGER_COLOURS.get(name, "#888")
        return (f'<div class="hl-card"><div class="hl-title">{desc[0]}</div>'
                f'<div class="hl-row"><div class="hl-dot" style="background:{color}"></div>'
                f'<div class="hl-name">{name}</div>'
                f'<div class="hl-val">{desc[1]}</div></div></div>')

    highlights_html = (
        hl_row(best_m,     ["Highest single race score", f"{best_pts} pts — {best_r}"]) +
        hl_row(worst_m,    ["Lowest single race score",  f"{worst_pts} pts — {worst_r}"]) +
        hl_row(best_avg_m["name"],  ["Best average per race",   f"{best_avg_m['avg']:.1f} pts/race"]) +
        hl_row(worst_avg_m["name"], ["Worst average per race",  f"{worst_avg_m['avg']:.1f} pts/race"]) +
        hl_row(cons_m["name"],  ["Most consistent",  cons_pos]) +
        hl_row(swing_m["name"], ["Biggest position swing", f"{swing_val} places"])
    )

    fin_grid = f"26px minmax(70px,1fr) repeat({N_FIN},26px) 36px"

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Stats · Season overview &amp; highlights</div>
<style>.srow{{display:grid;grid-template-columns:{fin_grid};align-items:center;gap:5px;padding:7px 0;border-bottom:0.5px solid #2a2a2a;min-width:420px}}</style>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">{total_race_scores}</div><div class="mc-lbl">Races scored so far</div></div>
  <div class="mc"><div class="mc-val">{chips_used_count}</div><div class="mc-lbl">Chips used league-wide</div></div>
  <div class="mc"><div class="mc-val">{cons_m['name']}</div><div class="mc-lbl">Most consistent ({cons_pos})</div></div>
  <div class="mc"><div class="mc-val">{swing_m['name']}</div><div class="mc-lbl">Biggest swing ({swing_val} places)</div></div>
</div>
<div class="section-label">Finish distribution</div>
<div class="card" style="padding:4px 16px;overflow-x:auto;-webkit-overflow-scrolling:touch">
  <div style="display:grid;grid-template-columns:{fin_grid};gap:5px;padding:6px 0 2px;min-width:420px">
    <div></div><div></div>{pos_headers}<div style="font-size:9px;color:#555;text-align:right">Pts</div>
  </div>
  {finish_rows_html}
</div>
<div class="section-label">Chip usage</div>
<div class="card">{chip_rows_html}</div>
<div class="section-label">Season highlights</div>
<div class="hl-grid">{highlights_html}</div>"""


# ── 07 Position Changes ───────────────────────────────────────────────────────
def panel_positions(data):
    M   = data["managers_sorted"]
    RD  = data["races_done"]
    ND  = data["n_done"]

    if ND == 0:
        return "<h1>🏎 The Undercut Collective</h1><div class='subtitle'>Position Changes · No races complete yet</div>"

    # Most gained / lost in a SINGLE race (not overall)
    best_single  = ("—", 0, "—")   # (manager, gain, race_name)
    worst_single = ("—", 0, "—")   # (manager, loss, race_name)
    for m in M:
        for ri in range(1, len(m["positions"])):
            prev = m["positions"][ri - 1]
            curr = m["positions"][ri]
            if prev is None or curr is None:
                continue
            change = prev - curr   # positive = moved up
            race_name = RD[ri]["name"] if ri < len(RD) else "—"
            if change > best_single[1]:
                best_single  = (m["name"], change, race_name)
            if change < worst_single[1]:
                worst_single = (m["name"], change, race_name)

    # Longest streak at current position
    streak_best = ("—", 0)   # (manager, streak_length)
    for m in M:
        if not m["positions"]:
            continue
        cur_pos = m["positions"][-1]
        if cur_pos is None:
            continue
        streak = 0
        for p in reversed(m["positions"]):
            if p == cur_pos:
                streak += 1
            else:
                break
        if streak > streak_best[1]:
            streak_best = (m["name"], streak)

    last_race_name = RD[-1]["name"] if RD else "—"

    # ── Standings movement table — clean card per manager ───────────────────
    # Sorted by current (latest) position
    sorted_cur = sorted(M, key=lambda m: (m["positions"][-1] if m["positions"] and m["positions"][-1] is not None else 99))

    # Header row: race name + round number above each badge column
    # Use first word of race name as abbreviation to keep it compact
    def short_race(name):
        # Use up to first 7 chars of first word, handles "Australia", "China", "Great Britain" → "Great"
        return name.split()[0][:7]

    header_badges = "".join(
        f'<div style="display:flex;flex-direction:column;align-items:center;gap:1px;width:28px;flex-shrink:0">'
        f'<div style="font-size:8px;color:#555;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:34px;text-align:center">{short_race(race["name"])}</div>'
        f'<div style="font-size:9px;color:#444">R{race["round"]}</div>'
        f'</div>'
        for race in RD
    )
    min_row_w = 80 + 34 * len(RD) + 80
    header_row = (
        f'<div style="display:flex;align-items:flex-end;gap:12px;padding:6px 0 4px;border-bottom:0.5px solid #2a2a2a;min-width:max(100%,{min_row_w}px)">'
        f'<div style="width:10px;flex-shrink:0"></div>'
        f'<div style="min-width:60px;flex-shrink:0"></div>'
        f'<div style="display:flex;gap:6px;flex:1">{header_badges}</div>'
        f'<div style="min-width:52px;text-align:right;flex-shrink:0;font-size:9px;color:#555;font-weight:500;text-transform:uppercase;letter-spacing:.04em">Last race</div>'
        f'</div>'
    )

    rows_html = ""
    for m in sorted_cur:
        if not m["positions"] or m["positions"][-1] is None:
            continue
        pos_cur   = m["positions"][-1]
        # Net = change since previous race (last two non-None positions)
        non_none  = [p for p in m["positions"] if p is not None]
        if len(non_none) >= 2:
            net = non_none[-2] - non_none[-1]   # positive = moved up
        else:
            net = 0

        # Net movement badge
        if net > 0:
            net_html = f'<span style="color:#1D9E75;font-size:13px;font-weight:600">▲ +{net}</span>'
        elif net < 0:
            net_html = f'<span style="color:#E24B4A;font-size:13px;font-weight:600">▼ {net}</span>'
        else:
            net_html = '<span style="color:#555;font-size:13px;font-weight:500">— 0</span>'

        # Race-by-race position badges (all races, skip Nones)
        race_badges = ""
        for ri, (race, pos) in enumerate(zip(RD, m["positions"])):
            if pos is None:
                continue
            is_cur = (ri == len(m["positions"]) - 1)
            bg = m["color"] if is_cur else "#1e1e1e"
            tc = "#0f0f0f" if is_cur else m["color"]
            race_badges += (
                f'<div style="display:flex;flex-direction:column;align-items:center;gap:2px">'
                f'<div style="width:28px;height:28px;border-radius:6px;background:{bg};border:1.5px solid {m["color"]};'
                f'display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;color:{tc}">P{pos}</div>'
                f'</div>'
            )

        rows_html += f"""<div style="display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:0.5px solid #2a2a2a;min-width:max(100%,{min_row_w}px)">
  <div style="width:10px;height:10px;border-radius:50%;background:{m['color']};flex-shrink:0"></div>
  <div style="min-width:60px;font-size:13px;font-weight:500;flex-shrink:0">{m['name']}</div>
  <div style="display:flex;gap:6px;flex:1">{race_badges}</div>
  <div style="min-width:52px;text-align:right;flex-shrink:0">{net_html}</div>
</div>"""

    # ── Position timeline chart ───────────────────────────────────────────────
    race_labels = [r["name"] for r in RD]
    datasets = []
    for i, m in enumerate(M):
        if not m["positions"]:
            continue
        # Convert None → null for JS; Chart.js will gap them with spanGaps:false
        # Use actual values only — Chart.js handles nulls fine with spanGaps:true
        pts = [p if p is not None else "null" for p in m["positions"]]
        pts_js = "[" + ",".join(str(p) for p in pts) + "]"
        dash = DASH_PATTERNS[i % len(DASH_PATTERNS)]
        datasets.append({
            "__name":       m["name"],
            "__pts":        pts_js,   # pre-rendered, not via js()
            "__color":      m["color"],
            "__dash":       js(dash),
        })

    legend_html = "".join(
        f'<span><span style="width:10px;height:10px;border-radius:2px;background:{m["color"]};display:inline-block;flex-shrink:0"></span>{m["name"]}</span>'
        for m in M)

    # Build datasets JS manually so pts_js goes in unquoted
    ds_js = "[" + ",".join(
        f'{{label:{js(d["__name"])},data:{d["__pts"]},borderColor:{js(d["__color"])},'
        f'borderDash:{d["__dash"]},borderWidth:2.5,pointBackgroundColor:{js(d["__color"])},'
        f'pointRadius:6,pointHoverRadius:8,fill:false,tension:0,spanGaps:true}}'
        for d in datasets
    ) + "]"

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Position Changes · How the standings have shifted</div>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">+{best_single[1]}</div><div class="mc-lbl">Biggest single-race climb — {best_single[0]} ({best_single[2]})</div></div>
  <div class="mc"><div class="mc-val">{worst_single[1]}</div><div class="mc-lbl">Biggest single-race drop — {worst_single[0]} ({worst_single[2]})</div></div>
  <div class="mc"><div class="mc-val">{streak_best[1]} race{'s' if streak_best[1] != 1 else ''}</div><div class="mc-lbl">Longest position streak — {streak_best[0]}</div></div>
  <div class="mc"><div class="mc-val">{last_race_name}</div><div class="mc-lbl">Most recent race</div></div>
</div>
<div class="section-label">Standings movement</div>
<div style="overflow-x:auto;-webkit-overflow-scrolling:touch"><div class="card" style="padding:4px 16px">
  {header_row}
  {rows_html}
  <div style="border-bottom:none;padding-bottom:4px"></div>
</div></div>
<div class="hint" style="margin-bottom:1.5rem">Filled badge = current position. Numbers show overall standings after each race.</div>
<div class="section-label">Position timeline</div>
<div style="position:relative;height:360px"><canvas id="posChart"></canvas></div>
<div class="legend" id="legend-positions"></div>
<script>
(function(){{
new Chart(document.getElementById('posChart'),{{
  type:'line',
  data:{{labels:{js(race_labels)},datasets:{ds_js}}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: P${{ctx.parsed.y}}`}}}}}},
    scales:{{
      x:{{ticks:{{color:'#888',font:{{size:12}}}},grid:{{color:'rgba(255,255,255,0.06)'}}}},
      y:{{reverse:true,min:0.5,max:{len(M)}+0.5,
        afterBuildTicks:function(ax){{ax.ticks=[...[{','.join(str(i) for i in range(1,len(M)+1))}].map(v=>{{return{{value:v}};}})]}},
        ticks:{{color:'#888',callback:function(v){{const r=Math.round(v);return(r>=1&&r<={len(M)})?'P'+r:'';}}}},
        grid:{{color:'rgba(255,255,255,0.06)'}}}}
    }}}}
}});
document.getElementById('legend-positions').innerHTML={js(legend_html)};
}})();
</script>"""


# ── 08 Team Picks ─────────────────────────────────────────────────────────────
def panel_picks(data):
    M           = data["managers_sorted"]
    RD          = data["races_done"]
    drv_results = data.get("driver_results", {})       # {driver_name: {race_name: pts}}
    con_results = data.get("constructor_results", {})  # {con_name:    {race_name: pts}}

    # Ordered list of races that have lineup data
    all_lineup_races = sorted(set(
        rname for m in M for rname in m["lineups"].keys()
    ), key=lambda r: next((rd["round"] for rd in data["races"] if rd["name"] == r), 99))

    if not all_lineup_races:
        return """<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Team Picks · No lineup data available yet</div>"""

    n_managers = len(M)

    # ── Season-level stats (across all races) ─────────────────────────────────
    driver_counts_all = {}
    con_counts_all    = {}
    drs_counts_all    = {}

    for m in M:
        for rname, picks in m["lineups"].items():
            for pick in picks:
                nm = pick["name"]
                if pick.get("is_constructor"):
                    con_counts_all[nm] = con_counts_all.get(nm, 0) + 1
                else:
                    driver_counts_all[nm] = driver_counts_all.get(nm, 0) + 1
                    if pick["drs"]:
                        drs_counts_all[nm] = drs_counts_all.get(nm, 0) + 1

    top_driver     = max(driver_counts_all, key=driver_counts_all.get) if driver_counts_all else "—"
    top_con        = max(con_counts_all,    key=con_counts_all.get)    if con_counts_all    else "—"
    top_drs        = max(drs_counts_all,    key=drs_counts_all.get)    if drs_counts_all    else "—"
    top_driver_cnt = driver_counts_all.get(top_driver, 0)
    top_con_cnt    = con_counts_all.get(top_con, 0)
    top_drs_cnt    = drs_counts_all.get(top_drs, 0)

    # ── Per-race JS data ──────────────────────────────────────────────────────
    # teams_by_race: per race, list of manager picks (with pts per pick, DRS flag)
    teams_by_race = {}
    for rname in all_lineup_races:
        teams_by_race[rname] = []
        for m in M:
            picks = m["lineups"].get(rname, [])
            if not picks:
                continue
            race_pts = sum((p["pts"] or 0) for p in picks)
            chip_by_race = m.get("chip_by_race", {})
            chip_name = chip_by_race.get(rname)
            chip_style = CHIP_STYLES.get(chip_name, {})
            teams_by_race[rname].append({
                "name":      m["name"],
                "teamName":  m["team"],
                "color":     m["color"],
                "racePts":   race_pts,
                "chip":      {"label": chip_name, "bg": chip_style.get("bg",""), "tc": chip_style.get("tc","")} if chip_name else None,
                "picks":     [{"name": p["name"], "drs": p["drs"], "drsMarker": p.get("drs_marker",""),
                               "pts": p["pts"], "isCon": p["is_constructor"]} for p in picks],
            })
        # Sort by race points descending so top scorer shows first
        teams_by_race[rname].sort(key=lambda t: -t["racePts"])

    # popularity_by_race: {race_name: {"drivers": [(name, count, pts), ...], "cons": [...]}}
    # Counts how many managers picked each driver/con that race, alongside their actual pts
    popularity_by_race = {}
    for rname in all_lineup_races:
        d_cnt, c_cnt = {}, {}
        for m in M:
            for pick in m["lineups"].get(rname, []):
                nm = pick["name"]
                if pick.get("is_constructor"):
                    c_cnt[nm] = c_cnt.get(nm, 0) + 1
                else:
                    d_cnt[nm] = d_cnt.get(nm, 0) + 1
        # Merge in actual points from results tables
        d_rows = sorted(d_cnt.items(), key=lambda x: -x[1])
        c_rows = sorted(c_cnt.items(), key=lambda x: -x[1])
        popularity_by_race[rname] = {
            "drivers": [{"name": n, "count": c,
                         "pts": (drv_results.get(n, {}).get(rname) or {}).get("pts")} for n, c in d_rows],
            "cons":    [{"name": n, "count": c,
                         "pts": (con_results.get(n, {}).get(rname) or {}).get("pts")} for n, c in c_rows],
            "maxD":    d_rows[0][1] if d_rows else 1,
            "maxC":    c_rows[0][1] if c_rows else 1,
        }

    # ── Trades per race ───────────────────────────────────────────────────────
    # Limitless races and the race immediately after are excluded from diffs —
    # the game auto-resets to the pre-Limitless lineup so neither transition
    # reflects a genuine transfer decision.
    trades_by_race = {}
    for i, rname in enumerate(all_lineup_races):
        if i == 0:
            continue
        prev_rname = all_lineup_races[i - 1]
        race_trades = []
        for m in M:
            cbr           = m.get("chip_by_race", {})
            is_wildcard   = cbr.get(rname) == "Wildcard"
            is_limitless  = cbr.get(rname) == "Limitless"
            was_limitless = cbr.get(prev_rname) == "Limitless"

            # Skip entirely for Limitless and post-Limitless — not real trades
            if is_limitless or was_limitless:
                continue

            prev_picks = m["lineups"].get(prev_rname, [])
            curr_picks = m["lineups"].get(rname, [])
            if not prev_picks or not curr_picks:
                continue

            prev_d = {p["name"] for p in prev_picks if not p["is_constructor"]}
            prev_c = {p["name"] for p in prev_picks if p["is_constructor"]}
            curr_d = {p["name"] for p in curr_picks if not p["is_constructor"]}
            curr_c = {p["name"] for p in curr_picks if p["is_constructor"]}

            sold_d   = sorted(prev_d - curr_d)
            bought_d = sorted(curr_d - prev_d)
            sold_c   = sorted(prev_c - curr_c)
            bought_c = sorted(curr_c - prev_c)

            paired = []
            for j in range(max(len(sold_d), len(bought_d))):
                paired.append({"out": sold_d[j]   if j < len(sold_d)   else None,
                                "in":  bought_d[j] if j < len(bought_d) else None,
                                "isCon": False})
            for j in range(max(len(sold_c), len(bought_c))):
                paired.append({"out": sold_c[j]   if j < len(sold_c)   else None,
                                "in":  bought_c[j] if j < len(bought_c) else None,
                                "isCon": True})

            total_trades = len(sold_d) + len(sold_c)

            if paired or is_wildcard:
                race_trades.append({
                    "name":       m["name"],
                    "color":      m["color"],
                    "team":       m["team"],
                    "wildcard":   is_wildcard,
                    "trades":     paired,
                    "tradeCount": total_trades,
                })

        race_trades.sort(key=lambda x: (-x["tradeCount"], x["name"]))
        trades_by_race[rname] = race_trades

    # ── DRS performance per manager ────────────────────────────────────────────
    # Collect ALL drs picks per race (Extra DRS chip gives both a 2X and 3X pick)
    drs_stats = []
    for m in M:
        races_with_drs = []
        for rname in all_lineup_races:
            picks = m["lineups"].get(rname, [])
            drs_picks = [p for p in picks if p["drs"]]
            if not drs_picks:
                continue
            for drs_pick in drs_picks:
                others = [p for p in picks if not p["drs"] and p["pts"] is not None]
                other_avg = round(sum(p["pts"] for p in others) / len(others), 1) if others else None
                races_with_drs.append({
                    "race":     rname,
                    "pick":     drs_pick["name"],
                    "pts":      drs_pick["pts"],
                    "marker":   drs_pick.get("drs_marker", "2X"),
                    "otherAvg": other_avg,
                })
        if not races_with_drs:
            continue
        scored = [r for r in races_with_drs if r["pts"] is not None]
        hit_rate   = round(sum(1 for r in scored if r["pts"] > 0) / len(scored) * 100) if scored else 0
        avg_pts    = round(sum(r["pts"] for r in scored) / len(scored), 1) if scored else 0
        drs_stats.append({
            "name":    m["name"],
            "color":   m["color"],
            "races":   races_with_drs,
            "hitRate": hit_rate,
            "avgPts":  avg_pts,
            "count":   len(races_with_drs),
        })
    drs_stats.sort(key=lambda x: -x["avgPts"])

    # ── Trade regrets ─────────────────────────────────────────────────────────
    # For each sold→bought pair, compute a combined Borda rank score across two
    # dimensions: points scored and budget change that race.
    # For each dimension, rank all assets of that type (driver or constructor) from
    # best to worst that race. Convert to a 0–1 percentile (1 = best in field).
    # Regret score = avg of sold_percentile − avg of bought_percentile across both dims.
    # A positive score means the sold asset was better-ranked than the replacement
    # on the combined pts+budget axis. Limitless races excluded.

    def build_rank_lookup(results_dict, race_name):
        """Return {asset_name: {"pts_pct": 0-1, "bud_pct": 0-1}} for a given race."""
        entries = []
        for name, races in results_dict.items():
            r = races.get(race_name, {}) or {}
            pts = r.get("pts")
            bud = r.get("bud")
            if pts is not None or bud is not None:
                entries.append((name, pts, bud))
        if not entries:
            return {}
        n = len(entries)
        # Rank by pts (higher = better)
        pts_sorted = sorted(entries, key=lambda x: (x[1] is not None, x[1] or 0))
        pts_rank   = {e[0]: i / (n - 1) if n > 1 else 0.5 for i, e in enumerate(pts_sorted)}
        # Rank by budget change (higher = better)
        bud_sorted = sorted(entries, key=lambda x: (x[2] is not None, x[2] or 0))
        bud_rank   = {e[0]: i / (n - 1) if n > 1 else 0.5 for i, e in enumerate(bud_sorted)}
        return {e[0]: {"pts_pct": pts_rank[e[0]], "bud_pct": bud_rank[e[0]]} for e in entries}

    regrets = []
    wins    = []
    for i, rname in enumerate(all_lineup_races[1:], 1):
        prev_rname = all_lineup_races[i - 1]

        # Build rank lookups for this race (drivers and constructors separately)
        drv_ranks = build_rank_lookup(drv_results, rname)
        con_ranks = build_rank_lookup(con_results, rname)

        for m in M:
            cbr = m.get("chip_by_race", {})
            if cbr.get(rname) == "Limitless" or cbr.get(prev_rname) == "Limitless":
                continue
            prev_picks = m["lineups"].get(prev_rname, [])
            curr_picks = m["lineups"].get(rname, [])
            if not prev_picks or not curr_picks:
                continue

            prev_d = {p["name"] for p in prev_picks if not p["is_constructor"]}
            prev_c = {p["name"] for p in prev_picks if p["is_constructor"]}
            curr_d = {p["name"] for p in curr_picks if not p["is_constructor"]}
            curr_c = {p["name"] for p in curr_picks if p["is_constructor"]}

            sold_d   = sorted(prev_d - curr_d)
            bought_d = sorted(curr_d - prev_d)
            sold_c   = sorted(prev_c - curr_c)
            bought_c = sorted(curr_c - prev_c)

            for pairs, is_con, results, ranks in [
                (list(zip(sold_d, bought_d)), False, drv_results, drv_ranks),
                (list(zip(sold_c, bought_c)), True,  con_results, con_ranks),
            ]:
                for sold_name, bought_name in pairs:
                    sold_r   = results.get(sold_name,   {}).get(rname) or {}
                    bought_r = results.get(bought_name, {}).get(rname) or {}
                    sold_pts   = sold_r.get("pts")
                    sold_bud   = sold_r.get("bud")
                    bought_pts = bought_r.get("pts")
                    bought_bud = bought_r.get("bud")
                    if sold_pts is None or bought_pts is None:
                        continue

                    sold_rank   = ranks.get(sold_name,   {})
                    bought_rank = ranks.get(bought_name, {})
                    if not sold_rank or not bought_rank:
                        continue

                    sold_score   = (sold_rank["pts_pct"]   + sold_rank["bud_pct"])   / 2
                    bought_score = (bought_rank["pts_pct"]  + bought_rank["bud_pct"]) / 2
                    diff_score   = sold_score - bought_score   # positive = regret, negative = good trade

                    entry = {
                        "manager":   m["name"],
                        "color":     m["color"],
                        "sold":      sold_name,
                        "bought":    bought_name,
                        "soldPts":   sold_pts,
                        "boughtPts": bought_pts,
                        "soldBud":   sold_bud,
                        "boughtBud": bought_bud,
                        "nextRace":  rname,
                        "soldAfter": prev_rname,
                        "isCon":     is_con,
                    }

                    if diff_score > 0 and sold_pts > bought_pts:
                        # Regret: sold outperformed bought
                        pts_margin = sold_pts - bought_pts
                        bud_margin = round((sold_bud or 0) - (bought_bud or 0), 2)
                        regrets.append({**entry,
                            "ptsMargin": pts_margin,
                            "budMargin": bud_margin,
                            "score":     round(diff_score * 100, 1),
                        })
                    elif diff_score < 0 and bought_pts > sold_pts:
                        # Win: bought outperformed sold
                        pts_margin = bought_pts - sold_pts
                        bud_margin = round((bought_bud or 0) - (sold_bud or 0), 2)
                        wins.append({**entry,
                            "ptsMargin": pts_margin,
                            "budMargin": bud_margin,
                            "score":     round(-diff_score * 100, 1),
                        })

    regrets.sort(key=lambda x: -x["score"])
    top_regrets = regrets[:3]
    wins.sort(key=lambda x: -x["score"])
    top_wins = wins[:3]

    # ── Loyalty ───────────────────────────────────────────────────────────────
    if len(all_lineup_races) >= 2:
        loyalty = {}
        for m in M:
            held = None
            for rname in all_lineup_races:
                names = {p["name"] for p in m["lineups"].get(rname, [])}
                held = names if held is None else held & names
            for asset in (held or set()):
                loyalty.setdefault(asset, []).append(m["name"])
        loyalty_highlights = sorted(
            [(asset, mgrs) for asset, mgrs in loyalty.items() if len(mgrs) >= 2],
            key=lambda x: -len(x[1]))[:4]
    else:
        loyalty_highlights = []

    # ── Trade count summary ───────────────────────────────────────────────────
    total_trades_per_manager = {}
    for rname, race_trades in trades_by_race.items():
        for entry in race_trades:
            n = entry["name"]
            total_trades_per_manager[n] = total_trades_per_manager.get(n, 0) + entry["tradeCount"]
    most_active     = max(total_trades_per_manager, key=total_trades_per_manager.get) if total_trades_per_manager else "—"
    most_active_cnt = total_trades_per_manager.get(most_active, 0)
    least_active    = min(total_trades_per_manager, key=total_trades_per_manager.get) if total_trades_per_manager else "—"
    least_active_cnt= total_trades_per_manager.get(least_active, 0)

    # ── HTML helpers ──────────────────────────────────────────────────────────
    _rd_names  = [r["name"] for r in data["races_done"]]
    _last_done = _rd_names[-1] if _rd_names else (all_lineup_races[-1] if all_lineup_races else "")
    trade_race_opts = "".join(
        f'<option value="{rn}"{" selected" if rn == _last_done else ""}>{rn}</option>'
        for rn in _rd_names[1:])
    race_options = "".join(
        f'<option value="{rn}"{" selected" if rn == _last_done else ""}>{rn}</option>'
        for rn in _rd_names)

    def fmt_pts(v):
        return f'+{v}' if v >= 0 else str(v)
    def fmt_bud(v):
        if v is None: return '—'
        return f'+{v:.1f}m' if v >= 0 else f'{v:.1f}m'

    loyalty_html = ""
    if loyalty_highlights:
        items = "".join(
            f'<div class="loyalty-item"><div class="loyalty-asset">{asset}</div>'
            f'<div class="loyalty-holders">{", ".join(mgrs)}</div></div>'
            for asset, mgrs in loyalty_highlights)
        loyalty_html = f"""
<div class="section-label">Loyal holds — picked every race</div>
<div class="card" style="padding:12px 16px">{items}</div>"""

    regret_html = ""
    if top_regrets:
        cards = ""
        for r in top_regrets:
            sold_pts_txt   = fmt_pts(r["soldPts"])
            bought_pts_txt = fmt_pts(r["boughtPts"])
            sold_bud_txt   = fmt_bud(r["soldBud"])
            bought_bud_txt = fmt_bud(r["boughtBud"])
            pts_margin_txt = fmt_pts(r["ptsMargin"])
            bud_margin_txt = fmt_bud(r["budMargin"])
            cards += (
                f'<div class="regret-card">'
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">'
                f'<div class="team-dot" style="background:{r["color"]}"></div>'
                f'<div class="team-name">{r["manager"]}</div>'
                f'<div style="margin-left:auto;font-size:11px;color:#888">{r["nextRace"]}</div>'
                f'</div>'
                f'<div class="regret-swap">'
                f'<div class="regret-asset regret-out">'
                f'<div class="regret-asset-name">{r["sold"]}</div>'
                f'<div class="regret-asset-pts" style="color:#4caf50">{sold_pts_txt} pts</div>'
                f'<div class="regret-asset-bud" style="color:#4caf50">{sold_bud_txt}</div>'
                f'</div>'
                f'<div class="regret-arrow">\u2192</div>'
                f'<div class="regret-asset regret-in">'
                f'<div class="regret-asset-name">{r["bought"]}</div>'
                f'<div class="regret-asset-pts" style="color:#f44336">{bought_pts_txt} pts</div>'
                f'<div class="regret-asset-bud" style="color:#f44336">{bought_bud_txt}</div>'
                f'</div>'
                f'</div>'
                f'<div class="regret-footer">'
                f'<span class="regret-margin">{pts_margin_txt} pts &middot; {bud_margin_txt} value</span>'
                f'<span class="regret-score">score {r["score"]}</span>'
                f'</div>'
                f'</div>'
            )
        regret_html = f"""
<div class="section-label">Trade regrets — sold too soon</div>
<div class="hint" style="margin-bottom:12px">Trades where the sold asset outranked its replacement on both points and value change the following race. Score is a combined percentile ranking (0–100). Limitless races excluded.</div>
<div class="regret-grid">{cards}</div>"""

    wins_html = ""
    if top_wins:
        cards = ""
        for r in top_wins:
            sold_pts_txt   = fmt_pts(r["soldPts"])
            bought_pts_txt = fmt_pts(r["boughtPts"])
            sold_bud_txt   = fmt_bud(r["soldBud"])
            bought_bud_txt = fmt_bud(r["boughtBud"])
            pts_margin_txt = fmt_pts(r["ptsMargin"])
            bud_margin_txt = fmt_bud(r["budMargin"])
            cards += (
                f'<div class="regret-card">'
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">'
                f'<div class="team-dot" style="background:{r["color"]}"></div>'
                f'<div class="team-name">{r["manager"]}</div>'
                f'<div style="margin-left:auto;font-size:11px;color:#888">{r["nextRace"]}</div>'
                f'</div>'
                f'<div class="regret-swap">'
                f'<div class="regret-asset" style="flex:1;background:#111;border-radius:6px;padding:7px 9px;border:0.5px solid #f4433644;min-width:0">'
                f'<div class="regret-asset-name">{r["sold"]}</div>'
                f'<div class="regret-asset-pts" style="color:#f44336">{sold_pts_txt} pts</div>'
                f'<div class="regret-asset-bud" style="color:#f44336">{sold_bud_txt}</div>'
                f'</div>'
                f'<div class="regret-arrow">\u2192</div>'
                f'<div class="regret-asset" style="flex:1;background:#111;border-radius:6px;padding:7px 9px;border:0.5px solid #4caf5044;min-width:0">'
                f'<div class="regret-asset-name">{r["bought"]}</div>'
                f'<div class="regret-asset-pts" style="color:#4caf50">{bought_pts_txt} pts</div>'
                f'<div class="regret-asset-bud" style="color:#4caf50">{bought_bud_txt}</div>'
                f'</div>'
                f'</div>'
                f'<div class="regret-footer">'
                f'<span style="font-size:11px;color:#4caf50;font-weight:500">+{r["ptsMargin"]} pts &middot; {bud_margin_txt} value</span>'
                f'<span class="regret-score">score {r["score"]}</span>'
                f'</div>'
                f'</div>'
            )
        wins_html = f"""
<div class="section-label">Best trades — great calls</div>
<div class="hint" style="margin-bottom:12px">Trades where the bought asset outranked what was sold on both points and value change the following race. Score is a combined percentile ranking (0–100). Limitless races excluded.</div>
<div class="regret-grid">{cards}</div>"""

    return f"""<h1>🏎 The Undercut Collective</h1>
<div class="subtitle">Team Picks · Lineups, trades &amp; transfer analysis</div>
<div class="metric-grid">
  <div class="mc"><div class="mc-val">{top_driver.split('. ')[-1] if '. ' in top_driver else top_driver}</div><div class="mc-lbl">Most picked driver ({top_driver_cnt}×)</div></div>
  <div class="mc"><div class="mc-val">{top_con}</div><div class="mc-lbl">Most picked constructor ({top_con_cnt}×)</div></div>
  <div class="mc"><div class="mc-val">{most_active.split(' ')[0] if most_active != '—' else '—'}</div><div class="mc-lbl">Most active trader ({most_active_cnt} moves)</div></div>
  <div class="mc"><div class="mc-val">{least_active.split(' ')[0] if least_active != '—' else '—'}</div><div class="mc-lbl">Most loyal ({least_active_cnt} moves)</div></div>
</div>

<div class="section-label">Lineup viewer</div>
<div style="display:flex;gap:12px;align-items:center;margin-bottom:1rem;flex-wrap:wrap">
  <label style="font-size:13px">Race</label>
  <select id="raceSel" onchange="ucPicksRace(this.value)">{race_options}</select>
  <label style="font-size:13px;margin-left:12px">Manager</label>
  <select id="teamSel" onchange="ucPicksTeam(raceSel.value, this.value)"></select>
</div>
<div id="picks-team-display"></div>

<div class="section-label">Driver &amp; constructor popularity</div>
<div class="hint" style="margin-bottom:12px">How many managers picked each driver/constructor for the selected race, with their actual points. Updates with the race selector above.</div>
<div class="pop-grid">
  <div>
    <div style="font-size:11px;color:#888;font-weight:500;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px">Drivers</div>
    <div class="card" style="padding:12px 16px"><div id="picks-driver-pop"></div></div>
  </div>
  <div>
    <div style="font-size:11px;color:#888;font-weight:500;text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px">Constructors</div>
    <div class="card" style="padding:12px 16px"><div id="picks-con-pop"></div></div>
  </div>
</div>

<div class="section-label">Transfers</div>
<div class="hint" style="margin-bottom:12px">Which managers made moves between races. Wildcard and Limitless weekends are flagged.</div>
<div style="display:flex;gap:12px;align-items:center;margin-bottom:1rem;flex-wrap:wrap">
  <label style="font-size:13px">Into race</label>
  <select id="tradeRaceSel">{trade_race_opts}</select>
</div>
<div id="picks-trades-display"></div>

<div class="section-label">DRS performance</div>
<div class="hint" style="margin-bottom:12px">Each manager's DRS pick results race by race. Extra DRS chip races show both picks separately. Hit rate = % of DRS picks that scored positive.</div>
<div id="picks-drs-display"></div>

{regret_html}
{wins_html}
{loyalty_html}
<script>
(function(){{
const teamsByRace={js(teams_by_race)};
const popByRace={js(popularity_by_race)};
const tradesByRace={js(trades_by_race)};
const drsStats={js(drs_stats)};
const nManagers={n_managers};
const p1Manager={js(M[0]["name"] if M else "")};
const raceSel=document.getElementById('raceSel');
const teamSel=document.getElementById('teamSel');
const tradeRaceSel=document.getElementById('tradeRaceSel');

/* ── Lineup viewer ── */
function showRace(rname){{
  const teams=teamsByRace[rname]||[];
  const prev=teamSel.value||p1Manager;
  teamSel.innerHTML='';
  teams.forEach(t=>{{
    const o=document.createElement('option');
    o.value=t.name;
    o.textContent=t.name+' \u2014 '+t.teamName;
    if(t.name===prev)o.selected=true;
    teamSel.appendChild(o);
  }});
  if(teams.length>0) showTeam(rname, teamSel.value);
  else document.getElementById('picks-team-display').innerHTML=
    '<div style="padding:1rem;color:#555;font-size:13px">No lineup data for this race yet.</div>';
  renderPop(rname);
}}

function showTeam(rname, manName){{
  const teams=teamsByRace[rname]||[];
  const t=teams.find(x=>x.name===manName);
  if(!t){{document.getElementById('picks-team-display').innerHTML='';return;}}
  const chipHtml=t.chip
    ?`<span class="chip-pill" style="background:${{t.chip.bg}};color:${{t.chip.tc}};margin-left:auto">${{t.chip.label}}</span>`:'';
  const drivers=t.picks.filter(p=>!p.isCon);
  const cons=t.picks.filter(p=>p.isCon);
  function pickCard(p){{
    const ptsTxt=p.pts!=null?(p.pts>=0?`+${{p.pts}}`:`${{p.pts}}`):'–';
    const ptsColor=p.pts==null?'#555':p.pts>0?'#4caf50':p.pts<0?'#f44336':'#888';
    const drsStyle=p.drs?`border-color:${{t.color}};background:${{t.color}}18`:'';
    const drsLabel=p.drsMarker?`<span style="font-size:10px;font-weight:600;color:${{t.color}};margin-left:4px">${{p.drsMarker}}</span>`:'';
    return `<div class="pick" style="${{drsStyle}}">
      <div class="pick-label">${{p.drs?'<span style="color:'+t.color+'">⚡ DRS</span>':p.isCon?'Constructor':'Driver'}}</div>
      <div class="pick-name">${{p.name}}${{drsLabel}}</div>
      <div class="pick-pts" style="color:${{ptsColor}}">${{ptsTxt}} pts</div>
    </div>`;
  }}
  const totalPts=t.racePts;
  const totalTxt=totalPts>=0?`+${{totalPts}}`:`${{totalPts}}`;
  const totalColor=totalPts>0?'#4caf50':totalPts<0?'#f44336':'#888';
  document.getElementById('picks-team-display').innerHTML=`
  <div class="team-card" style="margin-bottom:1rem">
    <div class="team-header">
      <div class="team-dot" style="background:${{t.color}}"></div>
      <div><div class="team-name">${{t.name}}</div><div class="team-sub">${{t.teamName}}</div></div>
      <div style="margin-left:auto;display:flex;align-items:center;gap:8px">
        ${{chipHtml}}
        <div style="font-size:18px;font-weight:600;color:${{totalColor}}">${{totalTxt}} pts</div>
      </div>
    </div>
    <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.05em;margin:10px 0 6px">Drivers</div>
    <div class="picks-grid">${{drivers.map(pickCard).join('')}}</div>
    <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.05em;margin:10px 0 6px">Constructors</div>
    <div class="picks-grid">${{cons.map(pickCard).join('')}}</div>
  </div>`;
}}

/* ── Trades ── */
function renderTrades(rname){{
  const entries=tradesByRace[rname]||[];
  const el=document.getElementById('picks-trades-display');
  if(!entries.length){{
    el.innerHTML='<div style="color:#555;font-size:13px;padding:8px 0">No transfers for this race.</div>';
    return;
  }}
  const cards=entries.map(e=>{{
    const wcBadge=e.wildcard
      ?`<span class="chip-pill" style="background:#4a1a12;color:#ff7a5a;margin-left:6px">Wildcard</span>`:'';
    const rows=e.trades.map(t=>{{
      const outHtml=t.out?`<span class="trade-out">\u2716 ${{t.out}}</span>`:`<span style="color:#333">\u2014</span>`;
      const inHtml =t.in ?`<span class="trade-in">\u2714 ${{t.in}}</span>` :`<span style="color:#333">\u2014</span>`;
      const type=t.isCon?`<span class="trade-type">CON</span>`:`<span class="trade-type">DRV</span>`;
      return `<div class="trade-row">${{type}}${{outHtml}}<span class="trade-arrow">\u2192</span>${{inHtml}}</div>`;
    }}).join('');
    const noTrades=!e.trades.length
      ?`<div style="font-size:12px;color:#555;margin-top:6px">No changes detected</div>`:'';
    return `<div class="trade-card">
      <div class="team-header" style="margin-bottom:${{e.trades.length?'10px':'4px'}}">
        <div class="team-dot" style="background:${{e.color}}"></div>
        <div style="flex:1"><div style="display:flex;align-items:center;flex-wrap:wrap;gap:2px"><span class="team-name">${{e.name}}</span>${{wcBadge}}</div><div class="team-sub">${{e.tradeCount}} transfer${{e.tradeCount!==1?'s':''}}</div></div>
      </div>
      ${{rows}}${{noTrades}}
    </div>`;
  }}).join('');
  el.innerHTML=`<div class="trade-grid">${{cards}}</div>`;
}}
tradeRaceSel.addEventListener('change', function(){{ renderTrades(this.value); }});

/* ── DRS table ── */
(function(){{
  const el=document.getElementById('picks-drs-display');
  if(!drsStats.length){{el.innerHTML='';return;}}
  const rows=drsStats.map(d=>{{
    const raceRows=d.races.map(r=>{{
      const ptsTxt=r.pts!=null?(r.pts>0?`<span style="color:#4caf50">+${{r.pts}}</span>`:`<span style="color:#f44336">${{r.pts}}</span>`):'–';
      const markerColor=r.marker==='3X'?'#88dd44':'#888';
      return `<div class="drs-race-row">
        <span class="drs-race-name">${{r.race}}</span>
        <span class="drs-pick-name">${{r.pick}} <span style="font-size:10px;color:${{markerColor}};font-weight:600">${{r.marker}}</span></span>
        <span class="drs-pts">${{ptsTxt}}</span>
      </div>`;
    }}).join('');
    const hitColor=d.hitRate>=75?'#4caf50':d.hitRate>=50?'#ffaa44':'#f44336';
    const avgTxt=d.avgPts>=0?`+${{d.avgPts}}`:`${{d.avgPts}}`;
    const avgColor=d.avgPts>0?'#4caf50':d.avgPts<0?'#f44336':'#888';
    return `<div class="drs-manager-card">
      <div class="team-header" style="margin-bottom:8px">
        <div class="team-dot" style="background:${{d.color}}"></div>
        <div class="team-name">${{d.name}}</div>
        <div style="margin-left:auto;display:flex;gap:16px;align-items:center">
          <div style="text-align:right"><div style="font-size:16px;font-weight:600;color:${{hitColor}}">${{d.hitRate}}%</div><div style="font-size:10px;color:#666">hit rate</div></div>
          <div style="text-align:right"><div style="font-size:16px;font-weight:600;color:${{avgColor}}">${{avgTxt}}</div><div style="font-size:10px;color:#666">avg pts</div></div>
        </div>
      </div>
      <div class="drs-races">${{raceRows}}</div>
    </div>`;
  }}).join('');
  el.innerHTML=rows;
}})();

/* ── Popularity bars ── */
function renderPop(rname){{
  const pop=popByRace[rname];
  if(!pop){{
    document.getElementById('picks-driver-pop').innerHTML='';
    document.getElementById('picks-con-pop').innerHTML='';
    return;
  }}
  function barRow(item, maxCnt, isDriver){{
    const pct=Math.round(item.count/maxCnt*100);
    const ptsStr=item.pts!=null?(item.pts>=0?`<span style="color:#4caf50">+${{item.pts}}</span>`:`<span style="color:#f44336">${{item.pts}}</span>`):'';
    const barColor=isDriver?'#378ADD':'#D85A30';
    return `<div class="pop-row">
      <div style="flex:1;min-width:0">
        <div style="display:flex;justify-content:space-between;align-items:baseline;gap:6px">
          <div class="pop-name">${{item.name}}</div>
          <div style="font-size:11px;color:#666;white-space:nowrap">${{ptsStr}}</div>
        </div>
        <div class="pop-bar-bg"><div class="pop-bar-fill" style="width:${{pct}}%;background:${{barColor}}"></div></div>
      </div>
      <div class="pop-count" style="white-space:nowrap">${{item.count}}/${{nManagers}}</div>
    </div>`;
  }}
  document.getElementById('picks-driver-pop').innerHTML=pop.drivers.map(d=>barRow(d,pop.maxD,true)).join('');
  document.getElementById('picks-con-pop').innerHTML=pop.cons.map(c=>barRow(c,pop.maxC,false)).join('');
}}

window.ucPicksRace=showRace;
window.ucPicksTeam=showTeam;
showRace(raceSel.value);
renderTrades(tradeRaceSel.value);
}})();
</script>"""




# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — ASSEMBLE COMBINED HTML
# ─────────────────────────────────────────────────────────────────────────────

SHARED_CSS = """
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #0f0f0f; color: #e8e8e6; min-height: 100vh; }
  .site-header { position: sticky; top: 0; z-index: 200; background: #0f0f0f; border-bottom: 1px solid #1e1e1e; }
  .header-inner { max-width: 900px; margin: 0 auto; padding: 0 1.5rem; }
  .header-top { display: flex; align-items: center; gap: 10px; padding: 10px 0 8px; border-bottom: 1px solid #1a1a1a; }
  .header-title { font-size: 15px; font-weight: 600; letter-spacing: -0.01em; }
  .header-badge { margin-left: auto; font-size: 10px; color: #666; background: #1a1a1a; border: 0.5px solid #2a2a2a; padding: 2px 8px; border-radius: 20px; }
  .tab-nav { display: flex; gap: 2px; overflow-x: auto; padding: 6px 0; scrollbar-width: none; }
  .tab-nav::-webkit-scrollbar { display: none; }
  .tab-btn { flex-shrink: 0; background: transparent; border: none; color: #555; font-size: 12px; font-weight: 500; font-family: inherit; padding: 5px 12px; border-radius: 6px; cursor: pointer; transition: color 0.12s, background 0.12s; white-space: nowrap; }
  .tab-btn:hover { color: #aaa; background: #1a1a1a; }
  .tab-btn.active { color: #e8e8e6; background: #1e1e1e; border: 0.5px solid #2a2a2a; }
  .tab-panel { display: none; }
  .tab-panel.active { display: block; }
  .panel-body { max-width: 860px; margin: 0 auto; padding: 1.5rem; }

  /* shared dashboard styles */
  h1 { font-size: 18px; font-weight: 600; margin-bottom: 4px; }
  .subtitle { font-size: 13px; color: #888; margin-bottom: 1.5rem; }
  .section-label { font-size: 11px; font-weight: 500; color: #888; letter-spacing: .05em; text-transform: uppercase; margin-bottom: 8px; margin-top: 1.5rem; }
  .metric-grid { display: grid; grid-template-columns: repeat(4,1fr); gap: 8px; margin-bottom: 1.5rem; }
  .mc { background: #1a1a1a; border-radius: 8px; padding: 12px 14px; border: 0.5px solid #2a2a2a; }
  .mc-val { font-size: 20px; font-weight: 500; }
  .mc-lbl { font-size: 11px; color: #888; margin-top: 2px; }
  .card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 4px 16px; margin-bottom: 1rem; }
  .hint { font-size: 11px; color: #555; margin-top: 6px; }
  .legend { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 10px; }
  .legend span { display: flex; align-items: center; gap: 4px; font-size: 12px; color: #888; }
  .chip-pill { display: inline-block; font-size: 10px; padding: 1px 7px; border-radius: 20px; font-weight: 500; margin-left: 4px; }
  .bar-bg { height: 4px; background: #2a2a2a; border-radius: 2px; overflow: hidden; margin-top: 5px; }
  .bar-fill { height: 100%; border-radius: 2px; }
  /* scrollable wrappers for wide tables */
  .scroll-x { overflow-x: auto; -webkit-overflow-scrolling: touch; }
  /* mobile */
  @media (max-width: 600px) {
    .panel-body { padding: 1rem; }
    .metric-grid { grid-template-columns: repeat(2,1fr); }
    .hl-grid { grid-template-columns: 1fr !important; }
    .mc-val { font-size: 16px; }
  }
  /* leaderboard */
  .podium-row { display: flex; gap: 6px; margin-bottom: 1.5rem; overflow-x: auto; -webkit-overflow-scrolling: touch; padding-bottom: 4px; }
  .podium-card { flex-shrink: 0; min-width: 160px; background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 8px; padding: 10px 12px; }
  .podium-pos { font-size: 11px; color: #888; margin-bottom: 2px; }
  .podium-name { font-size: 13px; font-weight: 500; }
  .row { display: grid; grid-template-columns: 26px 1fr 52px 52px; align-items: center; gap: 10px; padding: 10px 0; border-bottom: 0.5px solid #2a2a2a; }
  .row:last-child { border-bottom: none; }
  .pos-badge { width: 26px; height: 26px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 500; flex-shrink: 0; }
  .manager-name { font-size: 14px; font-weight: 500; line-height: 1.3; }
  .team-name { font-size: 11px; color: #888; }
  .gap-col { font-size: 11px; color: #888; text-align: right; white-space: nowrap; }
  .pts-col { font-size: 15px; font-weight: 500; text-align: right; white-space: nowrap; }
  /* race breakdown */
  .race-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 1rem 1.25rem; margin-bottom: 10px; }
  .race-header { display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 12px; }
  .race-title { font-size: 15px; font-weight: 500; }
  .race-round { font-size: 11px; color: #888; }
  .podium-strip { display: grid; grid-template-columns: repeat(3,1fr); gap: 6px; margin-bottom: 12px; }
  .slot { border-radius: 8px; padding: 8px 10px; display: flex; align-items: center; gap: 8px; }
  .medal { width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 10px; font-weight: 500; flex-shrink: 0; }
  .slot-name { font-size: 13px; font-weight: 500; }
  .slot-pts { font-size: 11px; color: #888; }
  .score-row { display: grid; grid-template-columns: 26px 1fr 44px; align-items: center; gap: 10px; padding: 5px 0; border-bottom: 0.5px solid #2a2a2a; }
  .score-row:last-child { border-bottom: none; }
  .pos-dot { width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 10px; font-weight: 500; flex-shrink: 0; }
  .score-name { font-size: 13px; font-weight: 500; }
  .score-pts { font-size: 13px; font-weight: 500; text-align: right; }
  .tag { display: inline-block; font-size: 11px; padding: 2px 8px; border-radius: 20px; background: #1e1e1e; color: #888; margin-right: 6px; margin-top: 8px; border: 0.5px solid #2a2a2a; }
  /* h2h */
  .vs-header { display: grid; grid-template-columns: 1fr auto 1fr; align-items: center; gap: 12px; margin-bottom: 1.5rem; }
  .vs-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 1rem 1.25rem; text-align: center; }
  .vs-card.winner { border-width: 2px; }
  .vs-name { font-size: 18px; font-weight: 500; }
  .vs-team { font-size: 11px; color: #888; margin-top: 2px; }
  .vs-pts { font-size: 28px; font-weight: 500; margin-top: 8px; }
  .vs-pos { font-size: 12px; color: #888; margin-top: 2px; }
  .vs-divider { font-size: 14px; font-weight: 500; color: #888; text-align: center; }
  .stat-grid { display: grid; grid-template-columns: 1fr auto 1fr; gap: 6px; margin-bottom: 1.5rem; }
  .stat-val { background: #1a1a1a; border-radius: 8px; padding: 8px 12px; font-size: 14px; font-weight: 500; text-align: center; border: 0.5px solid #2a2a2a; }
  .stat-val.hi { border-width: 1.5px; }
  .stat-label { background: #111; border-radius: 8px; padding: 8px 12px; font-size: 11px; color: #888; text-align: center; display: flex; align-items: center; justify-content: center; }
  .race-row { display: grid; grid-template-columns: 1fr 80px 1fr; align-items: center; gap: 8px; padding: 10px 0; border-bottom: 0.5px solid #2a2a2a; }
  .race-row:last-child { border-bottom: none; }
  .race-label { font-size: 12px; color: #888; text-align: center; }
  .chip-row { display: grid; grid-template-columns: 1fr 1fr 1fr; align-items: center; gap: 6px; padding: 8px 0; border-bottom: 0.5px solid #2a2a2a; }
  .chip-row:last-child { border-bottom: none; }
  .tick { width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; flex-shrink: 0; margin: 0 auto; }
  select { font-size: 13px; padding: 6px 10px; border-radius: 8px; border: 0.5px solid #333; background: #1a1a1a; color: #e8e8e6; cursor: pointer; }
  label { font-size: 13px; color: #888; }
  /* budget */
  .brow { display: grid; grid-template-columns: 26px 1fr 52px 60px 60px; align-items: center; gap: 10px; padding: 9px 0; border-bottom: 0.5px solid #2a2a2a; }
  .brow:last-child { border-bottom: none; }
  .dot { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }
  .bname { font-size: 13px; font-weight: 500; }
  .bcol { font-size: 13px; text-align: right; white-space: nowrap; }
  .col-hdr { font-size: 10px; color: #555; text-align: right; white-space: nowrap; }
  .pos { color: #1D9E75; } .neg { color: #E24B4A; } .neu { color: #888; }
  /* podiums */
  .race-podium { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 1rem 1.25rem; margin-bottom: 10px; }
  .podium-slots { display: grid; grid-template-columns: repeat(3,1fr); gap: 8px; }
  .tbd { font-size: 12px; color: #555; font-style: italic; }
  /* stats — .srow grid-template-columns injected dynamically by panel_stats */
  .srow:last-child { border-bottom: none; }
  .sname { font-size: 13px; font-weight: 500; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .fin { width: 24px; height: 24px; border-radius: 4px; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 500; }
  .f0 { background: #1e1e1e; color: #444; } .f1 { background: #FFD700; color: #7a5800; } .f2 { background: #C0C0C0; color: #4a4a4a; } .f3 { background: #CD7F32; color: #5a2d00; } .fn { background: #1e1e1e; color: #888; }
  .total-pts { font-size: 12px; font-weight: 500; text-align: right; }
  .hl-grid { display: grid; grid-template-columns: repeat(auto-fit,minmax(260px,1fr)); gap: 10px; margin-bottom: 1rem; }
  .hl-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 8px; padding: 10px 14px; }
  .hl-title { font-size: 12px; color: #888; margin-bottom: 6px; }
  .hl-row { display: flex; align-items: center; gap: 10px; }
  .hl-dot { width: 12px; height: 12px; border-radius: 50%; flex-shrink: 0; }
  .hl-name { font-size: 14px; font-weight: 500; }
  .hl-val { font-size: 13px; color: #888; }
  /* positions */
  .change-row { display: grid; align-items: center; gap: 10px; padding: 8px 0; border-bottom: 0.5px solid #2a2a2a; }
  .change-row:last-child { border-bottom: none; }
  .mname { font-size: 13px; font-weight: 500; }
  .pos-badge { width: 26px; height: 26px; border-radius: 6px; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: 500; background: #1e1e1e; color: #888; }
  .net-badge { font-size: 12px; font-weight: 500; text-align: right; }
  /* team picks */
  .team-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 1rem 1.25rem; margin-bottom: 10px; }
  .team-header { display: flex; align-items: center; gap: 10px; margin-bottom: 10px; }
  .team-dot { width: 12px; height: 12px; border-radius: 50%; flex-shrink: 0; }
  .team-name { font-size: 14px; font-weight: 500; }
  .team-sub { font-size: 11px; color: #888; }
  .picks-grid { display: grid; grid-template-columns: repeat(auto-fill,minmax(130px,1fr)); gap: 6px; }
  .pick { border-radius: 8px; padding: 7px 10px; background: #111; border: 0.5px solid #2a2a2a; }
  .pick-label { font-size: 9px; color: #555; text-transform: uppercase; letter-spacing: .04em; margin-bottom: 2px; }
  .pick-name { font-size: 12px; font-weight: 500; }
  .pick-pts { font-size: 12px; font-weight: 500; margin-top: 3px; }
  .drs-badge { display: inline-block; font-size: 9px; padding: 1px 5px; border-radius: 10px; background: #FFD700; color: #7a5800; font-weight: 500; margin-left: 4px; }
  /* trades */
  .trade-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(220px,1fr)); gap: 10px; }
  .trade-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 12px 14px; }
  .trade-row { display: flex; align-items: center; gap: 8px; padding: 4px 0; font-size: 12px; flex-wrap: wrap; }
  .trade-type { font-size: 9px; font-weight: 600; color: #555; text-transform: uppercase; letter-spacing: .04em; width: 28px; flex-shrink: 0; }
  .trade-out { color: #f44336; text-decoration: line-through; }
  .trade-in { color: #4caf50; }
  .trade-arrow { color: #444; flex-shrink: 0; }
  /* DRS performance */
  .drs-manager-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 12px 14px; margin-bottom: 10px; }
  .drs-races { border-top: 0.5px solid #2a2a2a; margin-top: 8px; padding-top: 6px; }
  .drs-race-row { display: grid; grid-template-columns: 90px 1fr 50px; align-items: center; gap: 8px; padding: 4px 0; border-bottom: 0.5px solid #1e1e1e; font-size: 12px; }
  .drs-race-row:last-child { border-bottom: none; }
  .drs-race-name { color: #888; font-size: 11px; }
  .drs-pick-name { font-weight: 500; }
  .drs-pts { text-align: right; font-weight: 500; }
  /* regrets */
  .regret-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(220px,1fr)); gap: 10px; margin-bottom: 1rem; }
  .regret-card { background: #1a1a1a; border: 0.5px solid #2a2a2a; border-radius: 10px; padding: 12px 14px; }
  .regret-swap { display: flex; align-items: stretch; gap: 8px; margin-bottom: 8px; }
  .regret-asset { flex: 1; background: #111; border-radius: 6px; padding: 7px 9px; min-width: 0; }
  .regret-out { border: 0.5px solid #4caf5044; }
  .regret-in  { border: 0.5px solid #f4433644; }
  .regret-asset-name { font-size: 12px; font-weight: 500; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; margin-bottom: 3px; }
  .regret-asset-pts { font-size: 13px; font-weight: 600; }
  .regret-asset-bud { font-size: 11px; margin-top: 2px; opacity: 0.85; }
  .regret-arrow { display: flex; align-items: center; color: #444; font-size: 14px; flex-shrink: 0; }
  .regret-footer { display: flex; justify-content: space-between; align-items: center; margin-top: 4px; }
  .regret-margin { font-size: 11px; color: #f44336; font-weight: 500; }
  .regret-score { font-size: 10px; color: #555; }
  /* loyalty */
  .loyalty-item { display: flex; justify-content: space-between; align-items: baseline; gap: 12px; padding: 7px 0; border-bottom: 0.5px solid #2a2a2a; font-size: 13px; }
  .loyalty-item:last-child { border-bottom: none; }
  .loyalty-asset { font-weight: 500; }
  .loyalty-holders { font-size: 11px; color: #888; text-align: right; }
  /* popularity */
  .pop-row { display: flex; align-items: center; gap: 8px; padding: 7px 0; border-bottom: 0.5px solid #2a2a2a; }
  .pop-row:last-child { border-bottom: none; }
  .pop-bar-bg { height: 4px; background: #2a2a2a; border-radius: 2px; overflow: hidden; margin-top: 3px; }
  .pop-bar-fill { height: 100%; border-radius: 2px; }
  .pop-name { font-size: 13px; font-weight: 500; }
  .pop-count { font-size: 12px; color: #888; text-align: right; white-space: nowrap; }
  .pop-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 1.5rem; }
"""

PANELS = [
    ("🏁 Leaderboard",    "leaderboard", panel_leaderboard),
    ("🗓 Race Breakdown",  "race",        panel_race_breakdown),
    ("⚔️ Head-to-Head",   "h2h",         panel_h2h),
    ("💰 Budget Tracker",  "budget",      panel_budget),
    ("🏆 Podiums",         "podiums",     panel_podiums),
    ("📊 Stats",           "stats",       panel_stats),
    ("📈 Positions",       "positions",   panel_positions),
    ("🧩 Team Picks",      "picks",       panel_picks),
]

def build_html(data):
    nd   = data["n_done"]
    nt   = data["n_total"]
    last = data["races_done"][-1]["name"] if data["races_done"] else "Pre-season"

    # NZ timestamp for "last updated" — tries zoneinfo (Python 3.9+), falls back to UTC+12/13
    try:
        from zoneinfo import ZoneInfo
        from datetime import datetime
        now_nz = datetime.now(ZoneInfo("Pacific/Auckland"))
    except Exception:
        from datetime import datetime, timezone, timedelta
        now_nz = datetime.now(timezone(timedelta(hours=12)))
    # Use cross-platform formatting (no %-d / %-I which are Linux-only)
    day  = str(now_nz.day)                          # no leading zero
    mon  = now_nz.strftime("%b")
    year = now_nz.strftime("%Y")
    hr   = now_nz.strftime("%I").lstrip("0") or "12"  # 12-hr no leading zero
    mn   = now_nz.strftime("%M")
    ampm = now_nz.strftime("%p").lower()
    updated_str = f"{day} {mon} {year}, {hr}:{mn} {ampm} NZT"

    tab_buttons = ""
    panel_divs  = ""
    for i, (label, slug, fn) in enumerate(PANELS):
        active = " active" if i == 0 else ""
        tab_buttons += f'      <button class="tab-btn{active}" id="tab-{slug}" onclick="showTab(\'{slug}\')">{label}</button>\n'
        try:
            content = fn(data)
        except Exception as e:
            content = f'<div style="padding:2rem;color:#E24B4A">Error generating panel: {e}</div>'
        panel_divs += f'  <div class="tab-panel{active}" id="panel-{slug}">\n    <div class="panel-body">\n{content}\n    </div>\n  </div>\n\n'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>The Undercut Collective — F1 Fantasy {SEASON}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
{SHARED_CSS}
</style>
</head>
<body>
<div class="site-header">
  <div class="header-inner">
    <div class="header-top">
      <span style="font-size:16px">🏎</span>
      <span class="header-title">The Undercut Collective</span>
      <span class="header-badge">{SEASON} Season · {nd}/{nt} races · Last: {last} · Updated {updated_str}</span>
    </div>
    <nav class="tab-nav">
{tab_buttons}    </nav>
  </div>
</div>
<div id="tab-content">
{panel_divs}</div>
<script>
function showTab(slug, fromRestore) {{
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('panel-' + slug).classList.add('active');
  document.getElementById('tab-' + slug).classList.add('active');
  document.getElementById('tab-' + slug).scrollIntoView({{behavior:'smooth',block:'nearest',inline:'center'}});
  window.scrollTo({{top:0, behavior: fromRestore ? 'instant' : 'smooth'}});
  try {{ sessionStorage.setItem('uc-tab', slug); }} catch(e) {{}}
}}
try {{
  const t = sessionStorage.getItem('uc-tab');
  if (t && document.getElementById('panel-' + t)) showTab(t, true);
}} catch(e) {{}}
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — GIT COMMIT + PUSH
# ─────────────────────────────────────────────────────────────────────────────

def git_push(last_race, repo_dir):
    if not GITHUB_REMOTE:
        print("Auto-push disabled (GITHUB_REMOTE is empty). Done.")
        return

    msg = f"{COMMIT_MSG_PREFIX} — {last_race}"
    try:
        def run(cmd):
            result = subprocess.run(cmd, cwd=repo_dir, capture_output=True, text=True)
            if result.returncode != 0:
                raise RuntimeError(result.stderr.strip() or result.stdout.strip())
            return result.stdout.strip()

        run(["git", "add", "index.html"])
        run(["git", "commit", "-m", msg])
        run(["git", "push", GITHUB_REMOTE])
        print(f"✅ Pushed to GitHub: '{msg}'")
    except FileNotFoundError:
        print("WARNING: git not found. Install Git from https://git-scm.com and try again.")
    except RuntimeError as e:
        err = str(e)
        if "nothing to commit" in err:
            print("No changes to commit (data unchanged since last push).")
        else:
            print(f"Git error: {err}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # Force UTF-8 output on Windows to avoid emoji encoding errors
    import sys, io
    if hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            pass
    print("=" * 55)
    print("  The Undercut Collective -- Dashboard Builder")
    print("=" * 55)

    raw  = read_workbook(WORKBOOK_PATH)
    data = compute(raw)

    nd   = data["n_done"]
    nt   = data["n_total"]
    last = data["races_done"][-1]["name"] if data["races_done"] else "Pre-season"
    print(f"  Season: {SEASON} | Races complete: {nd}/{nt} | Last: {last}")
    print(f"  Managers: {', '.join(m['name'] for m in data['managers_sorted'])}")
    print()

    html = build_html(data)
    OUTPUT_FILE.write_text(html, encoding="utf-8")
    size = OUTPUT_FILE.stat().st_size // 1024
    print(f"✅ Dashboard written: {OUTPUT_FILE} ({size} KB)")
    print()

    git_push(last, REPO_DIR)

    print()
    print("All done! Your friends can view the site at your GitHub Pages URL.")
    input("Press Enter to close...")


if __name__ == "__main__":
    main()
